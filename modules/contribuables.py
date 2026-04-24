"""modules/contribuables.py — Blueprint Contribuables (enrichi)"""
from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, session, send_file)
from datetime import datetime, date
from database import get_db
from modules.helpers import login_required, get_current_user, gen_num
import os, uuid

bp = Blueprint('contribuables', __name__)

# ─── Modules et leurs tables de référence ─────────────────────────
MODULES = [
    ('TNB',              'terrains',    'contribuable_id', '🏗️ TNB'),
    ('TDB',              'debits',      'contribuable_id', '🍺 TDB'),
    ('STATIONNEMENT',    'vehicules',   'contribuable_id', '🚗 Stat.'),
    ('FOURRIERE',        'dossiers_fourriere', 'contribuable_id', '🔑 Fourrière'),
    ('ODP',              'occupations', 'contribuable_id', '🎪 ODP'),
    ('LOCATION',         'locations',   'contribuable_id', '🏢 Location'),
    ('AFFERMAGE_SOUKS',  'affermages',  'contribuable_id', '🛒 Souks'),
]

MODULE_LABELS = {m[0]: m[3] for m in MODULES}


def _calc_impaye(conn, ctb_id):
    """Calcule le total impayé toutes déclarations émises non payées."""
    row = conn.execute(
        """SELECT COALESCE(SUM(d.montant_total),0) as s
           FROM declarations d
           LEFT JOIN bulletins b ON b.declaration_id=d.id
           WHERE d.contribuable_id=?
             AND d.statut NOT IN ('annule','paye')
             AND (b.id IS NULL OR b.statut NOT IN ('valide','encaisse','paye'))""",
        (ctb_id,)
    ).fetchone()
    return round(float(row['s'] or 0), 2)


def _get_modules_ctb(conn, ctb_id):
    """Retourne la liste des modules actifs pour un contribuable."""
    modules = []
    for mod, table, col, label in MODULES:
        try:
            n = conn.execute(
                f"SELECT COUNT(*) as c FROM {table} WHERE {col}=?", (ctb_id,)
            ).fetchone()['c']
            if n > 0:
                modules.append({'code': mod, 'label': label, 'nb': n})
        except Exception:
            pass  # table inexistante pour ce module
    return modules


# ════════════════════════════════════════════════════════════
#  LISTE
# ════════════════════════════════════════════════════════════
@bp.route('/contribuables')
@login_required
def contribuables():
    user = get_current_user()
    conn = get_db()
    q = request.args.get('q', '')
    filtre_module = request.args.get('module', '')

    sql = '''SELECT c.*, com.nom as commune_nom FROM contribuables c
        LEFT JOIN communes com ON c.commune_id=com.id WHERE c.actif=1'''
    params = []
    if q:
        sql += (' AND (c.nom LIKE ? OR c.prenom LIKE ? OR c.numero LIKE ?'
                ' OR c.raison_sociale LIKE ? OR c.cin LIKE ? OR c.nom_ar LIKE ?'
                ' OR c.telephone LIKE ? OR c.adresse LIKE ?)')
        params = [f'%{q}%'] * 8
    raw = conn.execute(sql + ' ORDER BY c.date_creation DESC', params).fetchall()

    # Enrichissement : total impayé + modules par contribuable
    items = []
    for row in raw:
        item = dict(row)
        item['total_impaye'] = _calc_impaye(conn, item['id'])
        item['modules']      = _get_modules_ctb(conn, item['id'])
        # Filtre par module (AUCUN = non affecté à aucun module)
        if filtre_module == 'AUCUN':
            if item['modules']:
                continue
        elif filtre_module:
            codes = [m['code'] for m in item['modules']]
            if filtre_module not in codes:
                continue
        items.append(item)

    communes = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    commune_logo = None
    c_row = conn.execute('SELECT logo FROM communes WHERE actif=1 LIMIT 1').fetchone()
    if c_row and c_row['logo']:
        commune_logo = c_row['logo']
    total_global_impaye = round(sum(i['total_impaye'] for i in items), 2)
    conn.close()
    return render_template('contribuables/contribuables.html',
                           user=user, items=items, communes=communes,
                           q=q, filtre_module=filtre_module,
                           total_global_impaye=total_global_impaye,
                           module_labels=MODULE_LABELS,
                           commune_logo=commune_logo,
                           today=date.today().isoformat())


@bp.route('/contribuables/export-excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    conn = get_db()
    data = conn.execute('''SELECT numero, nom, prenom, raison_sociale, cin, ice, telephone, email, adresse, type_personne, date_creation 
                           FROM contribuables WHERE actif=1 ORDER BY date_creation DESC''').fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Contribuables"
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1e3a5f')
    hdrs = ['N°', 'Nom', 'Prénom', 'Raison Sociale', 'CIN', 'ICE', 'Téléphone', 'Email', 'Adresse', 'Type', 'Date Enregistrement']
    
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hf
        cell.fill = hfill
        ws.column_dimensions[cell.column_letter].width = 18
        
    for r, row in enumerate(data, 2):
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'contribuables_{datetime.now():%Y%m%d}.xlsx')



# ════════════════════════════════════════════════════════════
#  DÉTAIL CONTRIBUABLE
# ════════════════════════════════════════════════════════════
@bp.route('/contribuables/<int:id>/detail')
@login_required
def ctb_detail(id):
    user = get_current_user()
    conn = get_db()
    contrib = conn.execute(
        '''SELECT c.*, com.nom as commune_nom
           FROM contribuables c LEFT JOIN communes com ON c.commune_id=com.id
           WHERE c.id=?''', (id,)
    ).fetchone()
    if not contrib:
        flash('Contribuable introuvable', 'danger')
        return redirect(url_for('contribuables.contribuables'))

    # Modules actifs
    modules = _get_modules_ctb(conn, id)

    # Déclarations par module
    declarations = conn.execute(
        '''SELECT d.*, b.statut as bull_statut, b.numero_bulletin, b.montant as bull_montant
           FROM declarations d
           LEFT JOIN bulletins b ON b.declaration_id=d.id
           WHERE d.contribuable_id=?
           ORDER BY d.date_creation DESC LIMIT 50''', (id,)
    ).fetchall()

    # Total impayé par module
    impaye_par_module = {}
    for mod, table, col, label in MODULES:
        try:
            row = conn.execute(
                """SELECT COALESCE(SUM(d.montant_total),0) as s
                   FROM declarations d
                   LEFT JOIN bulletins b ON b.declaration_id=d.id
                   WHERE d.contribuable_id=? AND d.module=?
                     AND d.statut NOT IN ('annule','paye')
                     AND (b.id IS NULL OR b.statut NOT IN ('valide','encaisse','paye'))""",
                (id, mod)
            ).fetchone()
            v = round(float(row['s'] or 0), 2)
            if v > 0:
                impaye_par_module[mod] = {'label': label, 'montant': v}
        except Exception:
            pass

    total_impaye = round(sum(v['montant'] for v in impaye_par_module.values()), 2)

    # Historique bulletins
    bulletins = conn.execute(
        '''SELECT b.*, d.module, d.annee, d.numero as decl_num
           FROM bulletins b
           LEFT JOIN declarations d ON d.id=b.declaration_id
           WHERE b.contribuable_id=?
           ORDER BY b.date_creation DESC LIMIT 30''', (id,)
    ).fetchall()

    total_encaisse = sum(
        float(b['montant'] or 0) for b in bulletins
        if b['statut'] in ('valide', 'encaisse', 'paye')
    )

    # Documents uploadés
    documents = conn.execute(
        'SELECT * FROM ctb_documents WHERE contribuable_id=? ORDER BY date_upload DESC', (id,)
    ).fetchall()

    conn.close()
    return render_template(
        'contribuables/ctb_detail.html',
        user=user, contrib=contrib,
        modules=modules,
        declarations=declarations,
        impaye_par_module=impaye_par_module,
        total_impaye=total_impaye,
        bulletins=bulletins,
        total_encaisse=total_encaisse,
        documents=documents,
        today=date.today().isoformat()
    )


# ════════════════════════════════════════════════════════════
#  UPLOAD DOCUMENT
# ════════════════════════════════════════════════════════════
@bp.route('/contribuables/<int:id>/upload', methods=['POST'])
@login_required
def ctb_upload(id):
    user = get_current_user()
    f = request.files.get('fichier')
    type_doc = request.form.get('type_doc', 'autre')
    if not f or not f.filename:
        flash('Aucun fichier sélectionné', 'warning')
        return redirect(url_for('contribuables.ctb_detail', id=id))
    ext = os.path.splitext(f.filename)[1].lower()
    fname = f"{uuid.uuid4().hex}{ext}"
    chemin = os.path.join('uploads', 'contribuables', fname)
    f.save(chemin)
    conn = get_db()
    conn.execute(
        '''INSERT INTO ctb_documents
           (contribuable_id, type_doc, nom_fichier, chemin, taille, date_upload, agent_id)
           VALUES (?,?,?,?,?,?,?)''',
        (id, type_doc, f.filename, chemin,
         os.path.getsize(chemin), date.today().isoformat(),
         user['id'] if user else None)
    )
    conn.commit(); conn.close()
    flash('Document uploadé ✅', 'success')
    return redirect(url_for('contribuables.ctb_detail', id=id))


@bp.route('/contribuables/doc/<int:doc_id>/supprimer', methods=['POST'])
@login_required
def ctb_doc_supprimer(doc_id):
    conn = get_db()
    doc = conn.execute('SELECT * FROM ctb_documents WHERE id=?', (doc_id,)).fetchone()
    ctb_id = 0
    if doc:
        ctb_id = doc['contribuable_id']
        try:
            if doc['chemin'] and os.path.exists(doc['chemin']):
                os.remove(doc['chemin'])
        except Exception:
            pass
        conn.execute('DELETE FROM ctb_documents WHERE id=?', (doc_id,))
        conn.commit()
    conn.close()
    flash('Document supprimé', 'info')
    return redirect(url_for('contribuables.ctb_detail', id=ctb_id))


@bp.route('/contribuables/doc/<int:doc_id>/telecharger')
@login_required
def ctb_doc_telecharger(doc_id):
    conn = get_db()
    doc = conn.execute('SELECT * FROM ctb_documents WHERE id=?', (doc_id,)).fetchone()
    conn.close()
    if doc and doc['chemin'] and os.path.exists(doc['chemin']):
        return send_file(doc['chemin'], as_attachment=True, download_name=doc['nom_fichier'])
    flash('Fichier introuvable', 'warning')
    return redirect(url_for('contribuables.contribuables'))


# ════════════════════════════════════════════════════════════
#  AVIS DE NON-PAIEMENT GLOBAL (tous modules)
# ════════════════════════════════════════════════════════════
@bp.route('/contribuables/<int:id>/avis-non-paiement')
@login_required
def ctb_avis_non_paiement(id):
    conn = get_db()
    contrib = conn.execute(
        '''SELECT c.*, com.nom as commune_nom
           FROM contribuables c LEFT JOIN communes com ON c.commune_id=com.id
           WHERE c.id=?''', (id,)
    ).fetchone()
    if not contrib:
        flash('Contribuable introuvable', 'danger')
        conn.close()
        return redirect(url_for('contribuables.contribuables'))

    commune_row = conn.execute('SELECT * FROM communes LIMIT 1').fetchone()
    commune  = commune_row['nom']      if commune_row else ''
    province = (commune_row['province']
                if commune_row and 'province' in commune_row.keys() else '')

    # Collecter impayés par module
    lignes = []
    for mod, table, col, label in MODULES:
        try:
            rows = conn.execute(
                """SELECT d.module, d.annee, d.trimestre, d.montant_total,
                          d.montant_principal, d.penalite_retard, d.majoration,
                          d.amende_non_declaration, d.date_echeance
                   FROM declarations d
                   LEFT JOIN bulletins b ON b.declaration_id=d.id
                   WHERE d.contribuable_id=? AND d.module=?
                     AND d.statut NOT IN ('annule','paye')
                     AND (b.id IS NULL OR b.statut NOT IN ('valide','encaisse','paye'))
                   ORDER BY d.annee, d.trimestre""",
                (id, mod)
            ).fetchall()
            for r in rows:
                lignes.append({
                    'module': mod,
                    'label':  label,
                    'annee':  r['annee'],
                    'trimestre': r['trimestre'],
                    'principal': float(r['montant_principal'] or 0),
                    'penalite':  float(r['penalite_retard'] or 0),
                    'majoration': float(r['majoration'] or 0),
                    'amende':    float(r['amende_non_declaration'] or 0),
                    'total':     float(r['montant_total'] or 0),
                    'echeance':  r['date_echeance'] or '',
                })
        except Exception:
            pass

    total_montant = round(sum(l['total'] for l in lignes), 2)

    # Grouper par module pour affichage
    from itertools import groupby
    lignes_g = {}
    for l in sorted(lignes, key=lambda x: x['module']):
        m = l['module']
        if m not in lignes_g:
            lignes_g[m] = {'label': l['label'], 'lignes': [], 'total': 0}
        lignes_g[m]['lignes'].append(l)
        lignes_g[m]['total'] += l['total']

    today_str = date.today().isoformat()
    n_avis = conn.execute(
        "SELECT COUNT(*) as c FROM avis_non_paiement"
    ).fetchone()['c'] + 1
    conn.close()
    avis_num = f"{n_avis}/{date.today().year}"

    return render_template('contribuables/ctb_avis_non_paiement.html',
        contrib=contrib, lignes_g=lignes_g, total_montant=total_montant,
        commune=commune, province=province,
        today=today_str, avis_num=avis_num,
        date_limite=f"{date.today().year}-12-31")


# ════════════════════════════════════════════════════════════
#  AJOUTER / MODIFIER / SUPPRIMER
# ════════════════════════════════════════════════════════════
@bp.route('/contribuables/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_contribuable():
    user = get_current_user()
    conn = get_db()
    communes = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    if request.method == 'POST':
        n = conn.execute('SELECT COUNT(*) as c FROM contribuables').fetchone()['c'] + 1
        num = f"CTB{datetime.now().year}{n:06d}"
        f = request.form
        conn.execute('''INSERT INTO contribuables
            (numero,type_personne,nom,prenom,nom_ar,prenom_ar,raison_sociale,raison_sociale_ar,
            cin,ice,rc,adresse,adresse_ar,ville,code_postal,telephone,email,commune_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (num, f.get('type_personne','physique'), f.get('nom',''), f.get('prenom',''),
             f.get('nom_ar',''), f.get('prenom_ar',''), f.get('raison_sociale',''), f.get('raison_sociale_ar',''),
             f.get('cin',''), f.get('ice',''), f.get('rc',''),
             f.get('adresse',''), f.get('adresse_ar',''), f.get('ville',''), f.get('code_postal',''),
             f.get('telephone',''), f.get('email',''), f.get('commune_id', 1)))
        conn.commit(); conn.close()
        flash('Contribuable ajouté ✅', 'success')
        return redirect(url_for('contribuables.contribuables'))
    conn.close()
    return render_template('contribuables/ajouter_contribuable.html', user=user, communes=communes)


@bp.route('/contribuables/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_contribuable(id):
    user = get_current_user()
    conn = get_db()
    contrib = conn.execute('SELECT * FROM contribuables WHERE id=?', (id,)).fetchone()
    communes = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE contribuables SET type_personne=?,nom=?,prenom=?,nom_ar=?,prenom_ar=?,
            raison_sociale=?,raison_sociale_ar=?,cin=?,ice=?,rc=?,adresse=?,adresse_ar=?,
            ville=?,code_postal=?,telephone=?,email=?,commune_id=? WHERE id=?''',
            (f.get('type_personne'), f.get('nom'), f.get('prenom'), f.get('nom_ar',''), f.get('prenom_ar',''),
             f.get('raison_sociale'), f.get('raison_sociale_ar',''), f.get('cin'), f.get('ice'), f.get('rc'),
             f.get('adresse'), f.get('adresse_ar',''), f.get('ville'), f.get('code_postal',''),
             f.get('telephone'), f.get('email'), f.get('commune_id'), id))
        conn.commit(); conn.close()
        flash('Contribuable modifié ✅', 'success')
        return redirect(url_for('contribuables.ctb_detail', id=id))
    conn.close()
    return render_template('contribuables/modifier_contribuable.html', user=user, contrib=contrib, communes=communes)


@bp.route('/contribuables/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_contribuable(id):
    user = get_current_user()
    if user['peut_supprimer']:
        conn = get_db()
        conn.execute('UPDATE contribuables SET actif=0 WHERE id=?', (id,))
        conn.commit(); conn.close()
    return redirect(url_for('contribuables.contribuables'))
