import pandas as pd
import re
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins

from dataclasses import dataclass

@dataclass
class Operation:
    num: int
    ref_paiement: int
    date: date
    partie_versante: str
    mode_paiement: str
    ref_creance: str
    code_rubrique: str
    libelle_rubrique: str
    principal: float
    majoration: float
    penalites: float
    astreintes: float
    total: float

    def to_dict(self):
        return {
            'num': self.num,
            'ref_paiement': self.ref_paiement,
            'date': self.date.isoformat() if self.date else None,
            'partie_versante': self.partie_versante,
            'mode_paiement': self.mode_paiement,
            'ref_creance': self.ref_creance,
            'code_rubrique': self.code_rubrique,
            'libelle_rubrique': self.libelle_rubrique,
            'principal': self.principal,
            'majoration': self.majoration,
            'penalites': self.penalites,
            'astreintes': self.astreintes,
            'total': self.total
        }

@dataclass
class MetaReleve:
    commune: str
    regie: str
    periode_debut: date
    periode_fin: date
    date_edition: date
    nb_pages: int
    total_principal: float
    total_majoration: float
    total_penalites: float
    total_astreintes: float
    total_general: float

    def to_dict(self):
        return {
            'commune': self.commune,
            'regie': self.regie,
            'periode_debut': self.periode_debut.isoformat() if self.periode_debut else None,
            'periode_fin': self.periode_fin.isoformat() if self.periode_fin else None,
            'date_edition': self.date_edition.isoformat() if self.date_edition else None,
            'nb_pages': self.nb_pages,
            'total_principal': self.total_principal,
            'total_majoration': self.total_majoration,
            'total_penalites': self.total_penalites,
            'total_astreintes': self.total_astreintes,
            'total_general': self.total_general
        }


def parse_releve(filepath: str) -> tuple[MetaReleve, list[Operation]]:
    engine = "xlrd" if filepath.endswith(".xls") else "openpyxl"
    df = pd.read_excel(filepath, header=None, engine=engine)
    
    meta = _extract_meta(df)
    header_row = _find_header_row(df)
    operations = _extract_operations(df, header_row + 1)
    
    total_p = 0.0
    total_m = 0.0
    total_pe = 0.0
    total_a = 0.0
    total_g = 0.0

    for i, op in enumerate(operations, 1):
        op.num = i
        total_p += op.principal
        total_m += op.majoration
        total_pe += op.penalites
        total_a += op.astreintes
        total_g += op.total
        
    meta.total_principal = total_p
    meta.total_majoration = total_m
    meta.total_penalites = total_pe
    meta.total_astreintes = total_a
    meta.total_general = total_g
    
    valid_dates = [op.date for op in operations if op.date]
    if valid_dates:
        meta.periode_debut = min(valid_dates)
        meta.periode_fin = max(valid_dates)
    
    return meta, operations


def _extract_meta(df) -> MetaReleve:
    commune = regie = periode = ""
    
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        for j, val in enumerate(row):
            v = str(val).strip().upper()
            if "COMMUNE/PROVINCE" in v:
                commune = str(df.iloc[i, min(j+1, len(row)-1)]).strip()
                if commune.lower() == 'nan' or not commune:
                    # try to find in remaining cols
                    for k in range(j+1, len(row)):
                        if str(row.iloc[k]).strip() and str(row.iloc[k]).strip().lower() != 'nan':
                            commune = str(row.iloc[k]).strip()
                            break
            elif "RÉGIE" in v or "REGIE" in v:
                regie = str(df.iloc[i, min(j+1, len(row)-1)]).strip()
                if regie.lower() == 'nan' or not regie:
                    for k in range(j+1, len(row)):
                        if str(row.iloc[k]).strip() and str(row.iloc[k]).strip().lower() != 'nan':
                            regie = str(row.iloc[k]).strip()
                            break
            elif "JOURNÉE" in v or "JOURNEE" in v:
                periode = str(df.iloc[i, min(j+1, len(row)-1)]).strip()
                if periode.lower() == 'nan' or not periode:
                    for k in range(j+1, len(row)):
                        if str(row.iloc[k]).strip() and str(row.iloc[k]).strip().lower() != 'nan':
                            periode = str(row.iloc[k]).strip()
                            break
                            
        if commune.lower() == 'nan' or commune.lower() == 'nat': commune = ""
        if regie.lower() == 'nan' or regie.lower() == 'nat': regie = ""
        if periode.lower() == 'nan' or periode.lower() == 'nat': periode = ""
        
    debut, fin = _parse_periode(periode)
    
    return MetaReleve(
        commune=commune,
        regie=regie,
        periode_debut=debut,
        periode_fin=fin,
        date_edition=date.today(),
        nb_pages=0,
        total_principal=0, total_majoration=0,
        total_penalites=0, total_astreintes=0,
        total_general=0
    )


def _find_header_row(df) -> int:
    for i in range(min(15, len(df))):
        for val in df.iloc[i]:
            if "MODE DE PAIEMENT" in str(val).upper():
                return i
    return 5


def _extract_operations(df, start_row: int) -> list[Operation]:
    ops = []
    for i in range(start_row, len(df)):
        row = df.iloc[i]
        
        if _is_total_row(row):
            continue # Skip totals, we recalculate them or we just want the rows. Wait, if it says TOTAL we break.
            
        # Stop on TOTAL
        is_total = False
        for val in row:
            if str(val).strip().upper().startswith("TOTAL"):
                is_total = True
                break
        if is_total:
            break
            
        ref_raw = row.iloc[3] if len(row) > 3 else None
        if pd.isna(ref_raw) or not str(ref_raw).replace('.0','').strip().isdigit():
            continue
            
        libelle_raw = str(row.iloc[12]) if len(row) > 12 else ""
        if libelle_raw == 'nan' and len(row) > 13:
             libelle_raw = str(row.iloc[13])
        code_rub, libelle_rub = _extract_libelle(libelle_raw)
        
        date_val = _parse_date(row.iloc[5]) if len(row) > 5 else None
        if not date_val and len(row) > 6:
            date_val = _parse_date(row.iloc[6])
            
        partie = str(row.iloc[7]).strip() if len(row) > 7 else ""
        if (partie.lower() == 'nan' or not partie) and len(row) > 8:
            partie = str(row.iloc[8]).strip()
        if partie.lower() == 'nan': partie = ""
            
        mode = str(row.iloc[1]).strip() if len(row) > 1 else ""
        if (mode.lower() == 'nan' or not mode) and len(row) > 2:
            mode = str(row.iloc[2]).strip()
        if mode.lower() == 'nan': mode = ""
            
        ref_cr = ""
        if len(row) > 10 and not pd.isna(row.iloc[10]):
            ref_cr = str(int(float(str(row.iloc[10]))))
        elif len(row) > 11 and not pd.isna(row.iloc[11]):
            try:
                ref_cr = str(int(float(str(row.iloc[11]))))
            except:
                ref_cr = str(row.iloc[11])
                
        # Indices in pandas could be slightly offset depending on the parse. Let's find columns by offset roughly.
        # usually 14, 16, 18, 21, 24
        # We will scan backwards from the end for the floats
        nums = []
        for v in row[::-1]:
            try:
                f = float(v)
                if not pd.isna(f):
                    nums.append(f)
            except:
                pass
        
        # the structure usually has total, astreintes, penalites, majoration, principal
        # so nums[0] is total, nums[1] is astreintes, etc.
        # But this is brittle. Let's stick to the indices in the prompt but allow slight variance
        prin = _safe_float(row.iloc[14]) if len(row) > 14 else 0.0
        if prin == 0.0 and len(row) > 15: prin = _safe_float(row.iloc[15])
        
        maj = _safe_float(row.iloc[16]) if len(row) > 16 else 0.0
        if maj == 0.0 and len(row) > 17: maj = _safe_float(row.iloc[17])
        
        pen = _safe_float(row.iloc[18]) if len(row) > 18 else 0.0
        if pen == 0.0 and len(row) > 19: pen = _safe_float(row.iloc[19])
        
        ast = _safe_float(row.iloc[21]) if len(row) > 21 else 0.0
        if ast == 0.0 and len(row) > 22: ast = _safe_float(row.iloc[22])
        
        tot = _safe_float(row.iloc[24]) if len(row) > 24 else 0.0
        if tot == 0.0 and len(row) > 25: tot = _safe_float(row.iloc[25])
        
        # Fallback if parsing totally fails but we have numbers
        if tot == 0.0 and nums:
            tot = nums[0]
            if len(nums) > 1: ast = nums[1]
            if len(nums) > 2: pen = nums[2]
            if len(nums) > 3: maj = nums[3]
            if len(nums) > 4: prin = nums[4]
            
        op = Operation(
            num=0,
            ref_paiement=int(float(str(ref_raw).replace('.0', ''))),
            date=date_val,
            partie_versante=partie,
            mode_paiement=mode,
            ref_creance=ref_cr,
            code_rubrique=code_rub,
            libelle_rubrique=libelle_rub,
            principal=prin,
            majoration=maj,
            penalites=pen,
            astreintes=ast,
            total=tot,
        )
        ops.append(op)
    return ops


def _extract_libelle(raw: str) -> tuple[str, str]:
    raw = str(raw).replace('\r\n', '\n').replace('\r', '\n')
    parts = raw.split('\n', 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    libelle = re.sub(r'^\s*\d+\s*', '', raw).strip()
    code_match = re.match(r'^\s*(\d+)', raw)
    code = code_match.group(1) if code_match else ""
    return code, libelle


def _is_total_row(row) -> bool:
    for val in row:
        v = str(val).upper().strip()
        if "SOUS TOTAL" in v:
            return True
    return False


def _safe_float(val) -> float:
    try:
        if pd.isna(val) or str(val).strip() == '':
            return 0.0
        return float(val)
    except:
        return 0.0


def _parse_date(val) -> date:
    try:
        if pd.isna(val): return None
        if hasattr(val, 'date'): return val.date()
        return pd.to_datetime(str(val)).date()
    except:
        return None


def _parse_periode(periode_str: str) -> tuple[date, date]:
    dates = re.findall(r'(\d{2}/\d{2}/\d{4})', periode_str)
    if len(dates) >= 2:
        fmt = "%d/%m/%Y"
        return datetime.strptime(dates[0], fmt).date(), datetime.strptime(dates[1], fmt).date()
    return None, None


# ── Lettre ──
UNITES = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf", "dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
DIZAINES = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante", "quatre-vingt", "quatre-vingt"]

def _centaines(n: int, pluriel_cent=True) -> str:
    if n == 0: return ""
    c, reste = divmod(n, 100)
    parts = []
    if c == 1: parts.append("cent")
    elif c > 1: parts.append(UNITES[c] + " cent" + ("s" if reste == 0 and pluriel_cent else ""))
    
    if reste > 0:
        if reste < 20: parts.append(UNITES[reste])
        else:
            d, u = divmod(reste, 10)
            if d == 7:
                parts.append("soixante et onze" if u == 1 else "soixante-" + UNITES[10 + u])
            elif d == 9:
                parts.append("quatre-vingt-" + UNITES[10 + u])
            elif d == 8:
                parts.append("quatre-vingts" if u == 0 else "quatre-vingt-" + UNITES[u])
            else:
                if u == 1 and d in (2, 3, 4, 5, 6): parts.append(DIZAINES[d] + " et un")
                elif u == 0: parts.append(DIZAINES[d])
                else: parts.append(DIZAINES[d] + "-" + UNITES[u])
    return " ".join(parts)

def chiffre_en_lettre(montant: float) -> str:
    if montant == 0: return "zéro Dirham"
    entier = int(montant)
    centimes = round((montant - entier) * 100)
    millions, reste = divmod(entier, 1_000_000)
    milliers, cents_part = divmod(reste, 1_000)
    parts = []
    if millions > 0:
        m_str = _centaines(millions)
        parts.append(m_str + (" million" if millions == 1 else " millions"))
    if milliers > 0:
        parts.append("mille" if milliers == 1 else _centaines(milliers) + " mille")
    if cents_part > 0:
        parts.append(_centaines(cents_part, pluriel_cent=(milliers > 0 or millions > 0)))
    total_str = " ".join(parts)
    result = total_str + f" Dirham{'s' if entier > 1 else ''}"
    if centimes > 0:
        result += f" et {_centaines(centimes)} centime{'s' if centimes > 1 else ''}"
    return result


# ── Export Excel ──
def exporter_xlsx(meta: MetaReleve, operations: list[Operation], filepath: str):
    BLEU_ENTETE  = "1F4E79"
    BLEU_CLAIR   = "DEEAF1"
    BLANC        = "FFFFFF"
    BLEU_COMMUNE = "2E75B6"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activité mensuelle"
    
    col_widths = {'A': 6, 'B': 16, 'C': 14, 'D': 40, 'E': 24, 'F': 14, 'I': 70}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
        
    ws.row_dimensions[1].height = 70
    
    def _write_info_row(r, label, value):
        ws.cell(row=r, column=3, value=label).font = Font(bold=True, color=BLEU_COMMUNE)
        ws.cell(row=r, column=4, value=value)
        
    _write_info_row(2, "Commune / Préfecture :", meta.commune)
    _write_info_row(3, "Régie :", meta.regie)
    per_str = f"Du {meta.periode_debut.strftime('%d/%m/%Y')} AU {meta.periode_fin.strftime('%d/%m/%Y')}" if meta.periode_debut and meta.periode_fin else "Période inconnue"
    _write_info_row(4, "Période :", per_str)
    
    ws['D5'] = "Total :"
    ws['D5'].font = Font(bold=True, color=BLEU_COMMUNE)
    ws['E5'] = f"=SUBTOTAL(9,Tableau_Ops[Total])"
    ws['E5'].number_format = '#,##0.00'
    ws['E5'].font = Font(bold=True, color=BLEU_COMMUNE)
    
    ws['A6'] = "*"
    ws['B6'] = chiffre_en_lettre(meta.total_general)
    ws['B6'].font = Font(italic=True, color="555555", size=9)
    ws.merge_cells('B6:F6')
    
    def _thin_border():
        s = Side(style='thin', color='B8CCE4')
        return Border(left=s, right=s, top=s, bottom=s)
        
    headers = ['N°', 'Réf. paiement', 'Date', 'Partie versante', 'Mode paiement', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=BLEU_ENTETE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
        
    ws['I7'] = "Libellé rubrique"
    ws['I7'].font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ws['I7'].fill = PatternFill("solid", fgColor=BLEU_ENTETE)
    ws['I7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 28
    
    for idx, op in enumerate(operations):
        row = 8 + idx
        is_pair = (idx % 2 == 1)
        bg = PatternFill("solid", fgColor=BLEU_CLAIR) if is_pair else PatternFill("solid", fgColor=BLANC)
        
        values = [op.num, op.ref_paiement, op.date, op.partie_versante, op.mode_paiement, op.total]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = bg
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical='center', horizontal='right' if col == 6 else 'center' if col in (1, 2, 3) else 'left')
            if col == 3 and val: cell.number_format = 'dd/mm/yyyy'
            if col == 6: cell.number_format = '#,##0.00'
            
        lib_cell = ws.cell(row=row, column=9, value=op.libelle_rubrique)
        lib_cell.fill = bg
        lib_cell.alignment = Alignment(vertical='center', wrap_text=True)
        
    last_data_row = 7 + len(operations)
    
    tbl_ref = f"A7:F{last_data_row}"
    tbl = Table(displayName="Tableau_Ops", ref=tbl_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    
    sig_row = last_data_row + 3
    ws.cell(row=sig_row, column=2, value="Fait à Aït Amira, le :").font = Font(italic=True)
    ws.cell(row=sig_row, column=3, value=date.today().strftime('%d/%m/%Y'))
    ws.cell(row=sig_row, column=5, value="Le Régisseur :").font = Font(bold=True)
    
    trait_row = sig_row + 3
    for col in range(4, 7):
        ws.cell(row=trait_row, column=col).border = Border(bottom=Side(style='thin', color="000000"))
        
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.44, right=0.25, top=0.45, bottom=0.35)
    ws.print_area = f"A1:I{sig_row + 4}"
    
    ws.oddHeader.center.text = f"&B COMMUNE AIT AMIRA – Activité mensuelle"
    ws.oddFooter.left.text = "Date : &D"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.print_options.horizontalCentered = True
    
    wb.save(filepath)
