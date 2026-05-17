"""
app.py — Orchestrateur principal (léger)
Toutes les routes métier sont dans modules/
"""
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, session, send_file, flash)
import sys, hashlib, io, json, os, shutil, threading, base64, logging, urllib.request as _urllib_req
from datetime import datetime, date, timedelta
from functools import wraps

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[logging.FileHandler('jibayat.log'), logging.StreamHandler()]
)
logger = logging.getLogger('jibayat')

# ── DB & helpers ────────────────────────────────────────────
from database import get_db, init_db
from modules.helpers import (login_required, get_current_user,
                              get_param, calculer_penalites, gen_num, annees_non_payees,
                              get_user_module_permissions, get_all_user_modules)

# ── Blueprints ───────────────────────────────────────────────
from modules.config         import bp as config_bp
from modules.contribuables  import bp as ctb_bp
from modules.tnb            import bp as tnb_bp
from modules.tdb            import bp as tdb_bp
from modules.stationnement  import bp as sta_bp
from modules.fourriere      import bp as fou_bp
from modules.occupation     import bp as odp_bp
from modules.location       import bp as loc_bp
from modules.souks          import bp as sou_bp
from modules.regie          import bp as regie_bp
from modules.emission       import bp as emission_bp
from modules.registre       import bp as registre_bp

# ── Application ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    # En mode PyInstaller, sys._MEIPASS pointe vers le dossier _internal
    base_dir = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    template_dir = os.path.join(base_dir, 'templates')
    static_dir = os.path.join(base_dir, 'static')
    app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
else:
    app = Flask(__name__)

# ── Clé secrète : env → config.json → fallback ──────────────
try:
    _sys_cfg = json.load(open('config.json', encoding='utf-8')) if os.path.exists('config.json') else {}
except Exception:
    _sys_cfg = {}
app.secret_key = os.environ.get('JIBAYAT_SECRET_KEY') or _sys_cfg.get('secret_key', 'super_secret_key_jibayat')

app.permanent_session_lifetime = timedelta(days=7)

# ── CSRF simplifié (sans Flask-WTF) ──────────────────────────
import secrets as _secrets

@app.before_request
def _csrf_check():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return
    # Routes exclues du CSRF
    if request.path.startswith('/api/') or request.path == '/login' or request.path == '/setup':
        return
    token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
    if not token or token != session.get('_csrf_token'):
        logger.warning(f"CSRF échec: {request.method} {request.path} depuis {request.remote_addr}")
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'ok': False, 'error': 'Token CSRF invalide'}), 403
        flash('Session expirée ou requête invalide. Veuillez réessayer.', 'danger')
        return redirect(url_for('index'))

@app.context_processor
def _inject_csrf():
    if '_csrf_token' not in session:
        session['_csrf_token'] = _secrets.token_hex(32)
    return {'csrf_token': session['_csrf_token']}

GITHUB_USER   = 'Yomix90'
GITHUB_REPO   = 'JIBAYAT'
GITHUB_BRANCH = 'main'

def _read_version():
    try:
        if os.path.exists('version.txt'):
            with open('version.txt', 'r', encoding='utf-8') as f:
                return f.read().strip()
    except Exception as e:
        logger.debug(f"Lecture version: {e}")
    return "1.0.0"

UPDATE_AVAILABLE = False

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    return f"<h1>Internal Server Error</h1><pre>{traceback.format_exc()}</pre>", 500

def _check_update_startup():
    global UPDATE_AVAILABLE
    try:
        import requests as _req, json
        cfg = {}
        if os.path.exists('config.json'):
            with open('config.json', 'r', encoding='utf-8') as f:
                cfg = json.load(f)
        token = cfg.get('github_token', '').strip()
        
        url = f'https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/version.txt'
        headers = {'Accept': 'application/vnd.github.v3.raw'}
        if token:
            headers['Authorization'] = f'token {token}'
            
        r = _req.get(url, headers=headers, timeout=10)
        
        if r.status_code == 200:
            remote = r.text.strip()
            local  = _read_version()
            def vt(v):
                try: return tuple(int(x) for x in v.split('.'))
                except: return (0,)
            if vt(remote) > vt(local):
                UPDATE_AVAILABLE = True
    except Exception as e:
        logger.debug(f"Vérification mise à jour: {e}")

# Lancement de la vérification en arrière-plan sans bloquer le serveur
threading.Thread(target=_check_update_startup, daemon=True).start()

@app.context_processor
def inject_global_vars():
    """Inject variables into all Jinja templates automatically."""
    user = get_current_user()
    user_modules = get_all_user_modules(user) if user else {}
    
    return {
        'sys_version': _read_version(),
        'sys_has_update': UPDATE_AVAILABLE,
        'user_modules': user_modules,
    }

for bp in (config_bp, ctb_bp, tnb_bp, tdb_bp, sta_bp, fou_bp, odp_bp, loc_bp, sou_bp, regie_bp):
    app.register_blueprint(bp)

app.register_blueprint(emission_bp, url_prefix='/emission')
app.register_blueprint(registre_bp)

DB = 'fiscalite.db'



# ── Contexte global (badge sidebar bulletins en attente) ─────────
@app.context_processor
def inject_global_counts():
    """Injecte nb_attente dans tous les templates pour le badge sidebar."""
    try:
        if 'user_id' not in session:
            return {'nb_attente': 0}
        conn = get_db()
        nb = conn.execute(
            "SELECT COUNT(*) as c FROM bulletins WHERE statut='en_attente'"
        ).fetchone()['c']
        conn.close()
        return {'nb_attente': nb}
    except Exception:
        return {'nb_attente': 0}

# ════════════════════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════════════════════
@app.route('/login', methods=['GET', 'POST'])
def login():
    conn = get_db()
    commune = conn.execute('SELECT nom, logo FROM communes WHERE id=1').fetchone()
    commune_nom = commune['nom'] if commune else 'COMMUNE'
    
    if request.method == 'POST':
        pwd = hashlib.sha256(request.form['password'].encode()).hexdigest()
        user = conn.execute('SELECT * FROM utilisateurs WHERE email=? AND mot_de_passe=? AND actif=1',
                            (request.form['email'], pwd)).fetchone()
        conn.close()
        if user:
            session['user_id'] = user['id']
            return redirect(url_for('index'))
        return render_template('login.html', error='Email ou mot de passe incorrect', commune_nom=commune_nom)
        
    conn.close()
    return render_template('login.html', commune_nom=commune_nom)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ════════════════════════════════════════════════════════════
#  DASHBOARD
# ════════════════════════════════════════════════════════════
@app.route('/')
@login_required
def index():
    user = get_current_user()
    conn = get_db()
    annee_cur = datetime.now().year

    def q(sql, *args):
        row = conn.execute(sql, args).fetchone()
        return row[0] if row else 0

    # ── Stats globales ────────────────────────────────────────
    stats = {
        'contribuables':     q('SELECT COUNT(*) FROM contribuables WHERE actif=1'),
        'bulletins_attente': q("SELECT COUNT(*) FROM bulletins WHERE statut='en_attente'"),
        'avis_emis':         q("SELECT COUNT(*) FROM avis_non_paiement WHERE statut='emis'"),
        'total_emis':        q('SELECT COALESCE(SUM(montant_total),0) FROM declarations'),
        'total_paye':        q("SELECT COALESCE(SUM(montant),0) FROM bulletins WHERE statut='paye'"),
    }

    # ── Stats par module ──────────────────────────────────────
    modules_stats = {
        'TNB': {
            'label': 'TNB — Terrains', 'icon': '🏗️', 'color': '#e8a020',
            'count': q('SELECT COUNT(*) FROM tnb_terrains'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='TNB' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='TNB' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'tnb.tnb_liste',
        },
        'TDB': {
            'label': 'Débits Boissons', 'icon': '🍺', 'color': '#8e44ad',
            'count': q('SELECT COUNT(*) FROM tdb_etablissements'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='DEBITS_BOISSONS' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='DEBITS_BOISSONS' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'tdb.tdb_liste',
        },
        'STA': {
            'label': 'Stationnement', 'icon': '🚗', 'color': '#2980b9',
            'count': q('SELECT COUNT(*) FROM sta_vehicules'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='STATIONNEMENT' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='STATIONNEMENT' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'sta.sta_liste',
        },
        'FOU': {
            'label': 'Fourrière', 'icon': '🔑', 'color': '#c0392b',
            'count': q('SELECT COUNT(*) FROM fou_dossiers'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='FOURRIERE' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='FOURRIERE' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'fou.fou_liste',
        },
        'ODP': {
            'label': 'Domaine Public', 'icon': '🏪', 'color': '#16a085',
            'count': q('SELECT COUNT(*) FROM odp_occupations'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='OCCUPATION_DOMAINE' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='OCCUPATION_DOMAINE' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'odp.odp_liste',
        },
        'LOC': {
            'label': 'Location Locaux', 'icon': '🏢', 'color': '#27ae60',
            'count': q('SELECT COUNT(*) FROM loc_locaux'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='LOCATION_LOCAUX' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='LOCATION_LOCAUX' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'loc.loc_liste',
        },
        'SOU': {
            'label': 'Souks', 'icon': '🛒', 'color': '#d35400',
            'count': q('SELECT COUNT(*) FROM sou_contrats'),
            'emis':  q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module='AFFERMAGE_SOUKS' AND annee=?", annee_cur),
            'paye':  q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module='AFFERMAGE_SOUKS' AND b.statut='paye' AND d.annee=?", annee_cur),
            'url': 'sou.sou_liste',
        },
    }

    # ── Recouvrement par mois (année courante) ────────────────
    mois_labels = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    mois_paye = []
    mois_emis = []
    for m in range(1, 13):
        mois_paye.append(q(
            "SELECT COALESCE(SUM(b.montant),0) FROM bulletins b WHERE b.statut='paye' AND strftime('%Y',b.date_paiement)=? AND strftime('%m',b.date_paiement)=?",
            str(annee_cur), f'{m:02d}'
        ))
        mois_emis.append(q(
            "SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE strftime('%Y',date_creation)=? AND strftime('%m',date_creation)=? AND statut='emis'",
            str(annee_cur), f'{m:02d}'
        ))

    # ── Répartition recettes par module ──────────────────────
    chart_modules_labels = [v['label'] for v in modules_stats.values()]
    chart_modules_data   = [v['paye'] for v in modules_stats.values()]
    chart_modules_colors = [v['color'] for v in modules_stats.values()]

    recentes = conn.execute('''SELECT d.*, c.nom, c.prenom, c.raison_sociale FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id ORDER BY d.date_creation DESC LIMIT 10''').fetchall()
    conn.close()

    return render_template('index.html', user=user, stats=stats,
                           modules_stats=modules_stats, recentes=recentes,
                           annee_cur=annee_cur,
                           mois_labels=mois_labels,
                           mois_paye=mois_paye,
                           mois_emis=mois_emis,
                           chart_modules_labels=chart_modules_labels,
                           chart_modules_data=chart_modules_data,
                           chart_modules_colors=chart_modules_colors)


# ════════════════════════════════════════════════════════════
#  DASHBOARD PAR MODULE
# ════════════════════════════════════════════════════════════

# Config de chaque module
MOD_DASHBOARD_CONFIG = {
    'tnb': {
        'label': 'TNB — Terrains Non Bâtis', 'icon': '🏗️', 'color': '#e8a020',
        'module_key': 'TNB', 'table': 'tnb_terrains',
        'liste_url_func': 'tnb.tnb_liste',
    },
    'tdb': {
        'label': 'Débits de Boissons', 'icon': '🍺', 'color': '#8e44ad',
        'module_key': 'DEBITS_BOISSONS', 'table': 'tdb_etablissements',
        'liste_url_func': 'tdb.tdb_liste',
    },
    'sta': {
        'label': 'Stationnement TPV', 'icon': '🚗', 'color': '#2980b9',
        'module_key': 'STATIONNEMENT', 'table': 'sta_vehicules',
        'liste_url_func': 'sta.sta_liste',
    },
    'fou': {
        'label': 'Fourrière', 'icon': '🔑', 'color': '#c0392b',
        'module_key': 'FOURRIERE', 'table': 'fou_dossiers',
        'liste_url_func': 'fou.fou_liste',
    },
    'odp': {
        'label': 'Occupation Domaine Public', 'icon': '🎪', 'color': '#16a085',
        'module_key': 'OCCUPATION_DOMAINE', 'table': 'odp_occupations',
        'liste_url_func': 'odp.odp_liste',
    },
    'loc': {
        'label': 'Location Locaux Commerciaux', 'icon': '🏢', 'color': '#27ae60',
        'module_key': 'LOCATION_LOCAUX', 'table': 'loc_locaux',
        'liste_url_func': 'loc.loc_liste',
    },
    'sou': {
        'label': 'Souks Communaux', 'icon': '🛒', 'color': '#d35400',
        'module_key': 'AFFERMAGE_SOUKS', 'table': 'sou_contrats',
        'liste_url_func': 'sou.sou_liste',
    },
}

@app.route('/module/<module>/dashboard')
@login_required
def mod_dashboard(module):
    if module not in MOD_DASHBOARD_CONFIG:
        flash('Module inconnu', 'danger')
        return redirect(url_for('index'))

    user = get_current_user()
    cfg  = MOD_DASHBOARD_CONFIG[module]
    mkey = cfg['module_key']
    annee_cur = datetime.now().year
    conn = get_db()

    def q(sql, *args):
        row = conn.execute(sql, args).fetchone()
        return row[0] if row else 0

    # ── KPI principaux ────────────────────────────────────
    total_dossiers = q(f'SELECT COUNT(*) FROM {cfg["table"]}')
    nb_decl_annee  = q("SELECT COUNT(*) FROM declarations WHERE module=? AND annee=?", mkey, annee_cur)
    nb_emis        = q("SELECT COUNT(*) FROM declarations WHERE module=? AND annee=? AND statut='emis'", mkey, annee_cur)
    nb_paye        = q("SELECT COUNT(*) FROM declarations WHERE module=? AND statut='paye'", mkey)
    nb_impaye      = q("SELECT COUNT(*) FROM declarations WHERE module=? AND statut='emis'", mkey)
    nb_annule      = q("SELECT COUNT(*) FROM declarations WHERE module=? AND statut='annule'", mkey)
    nb_sous_seuil  = q("SELECT COUNT(*) FROM declarations WHERE module=? AND annee=? AND statut='sous_seuil'", mkey, annee_cur)
    nb_avis        = q("SELECT COUNT(*) FROM avis_non_paiement a JOIN declarations d ON d.id=a.declaration_id WHERE d.module=? AND a.statut='emis'", mkey)

    total_emis  = q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module=? AND statut IN ('emis','paye')", mkey)
    total_paye  = q("SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module=? AND b.statut='paye'", mkey)
    total_impaye = q("SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module=? AND statut='emis'", mkey)
    taux = round(total_paye / total_emis * 100, 1) if total_emis > 0 else 0.0

    stats = {
        'total_dossiers': total_dossiers,
        'nb_decl_annee': nb_decl_annee,
        'nb_emis': nb_emis,
        'nb_paye': nb_paye,
        'nb_impaye': nb_impaye,
        'nb_annule': nb_annule,
        'nb_sous_seuil': nb_sous_seuil,
        'nb_avis': nb_avis,
        'total_emis': round(total_emis, 2),
        'total_paye': round(total_paye, 2),
        'total_impaye': round(total_impaye, 2),
        'taux_recouvrement': taux,
    }

    # ── Mensuel ───────────────────────────────────────────
    mois_labels = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    mois_paye, mois_emis_m = [], []
    for m in range(1, 13):
        mois_paye.append(q(
            "SELECT COALESCE(SUM(b.montant),0) FROM bulletins b JOIN declarations d ON b.declaration_id=d.id WHERE d.module=? AND b.statut='paye' AND strftime('%Y',b.date_paiement)=? AND strftime('%m',b.date_paiement)=?",
            mkey, str(annee_cur), f'{m:02d}'
        ))
        mois_emis_m.append(q(
            "SELECT COALESCE(SUM(montant_total),0) FROM declarations WHERE module=? AND strftime('%Y',date_creation)=? AND strftime('%m',date_creation)=? AND statut IN ('emis','paye')",
            mkey, str(annee_cur), f'{m:02d}'
        ))

    # ── Par année ─────────────────────────────────────────
    stats_par_annee = conn.execute("""
        SELECT d.annee,
               COUNT(*) as nb,
               COALESCE(SUM(CASE WHEN d.statut IN ('emis','paye') THEN d.montant_total ELSE 0 END),0) as total_emis,
               COALESCE(SUM(CASE WHEN d.statut='paye' THEN d.montant_total ELSE 0 END),0) as total_paye
        FROM declarations d
        WHERE d.module=?
        GROUP BY d.annee ORDER BY d.annee DESC LIMIT 8
    """, (mkey,)).fetchall()

    # ── Déclarations récentes ─────────────────────────────
    recentes = conn.execute("""
        SELECT d.*, c.nom, c.prenom, c.raison_sociale
        FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.module=?
        ORDER BY d.date_creation DESC LIMIT 8
    """, (mkey,)).fetchall()

    conn.close()

    mod_info = dict(cfg)
    mod_info['liste_url'] = url_for(cfg['liste_url_func'])

    return render_template('module_dashboard.html',
        user=user, module=module,
        mod_info=mod_info, stats=stats,
        annee_cur=annee_cur,
        mois_labels=mois_labels,
        mois_paye=mois_paye,
        mois_emis=mois_emis_m,
        stats_par_annee=[dict(r) for r in stats_par_annee],
        recentes=recentes)


# ════════════════════════════════════════════════════════════
#  DÉCLARATIONS (communes à tous les modules)
# ════════════════════════════════════════════════════════════
@app.route('/declarations/creer', methods=['POST'])
@login_required
def creer_declaration():
    user = get_current_user()
    f = request.form
    module     = f['module']
    ref_id     = int(f['reference_id'])
    contrib_id = int(f['contribuable_id'])
    annee      = int(f.get('annee', datetime.now().year))
    base       = float(f.get('base_calcul', 0))
    taux       = float(f.get('taux', 0))
    principal  = round(base * taux / 100 if taux else base, 2)
    date_ech   = f.get('date_echeance', '')
    date_decl  = f.get('date_declaration', date.today().isoformat())
    hors_delai = f.get('hors_delai') == '1'
    penalite, majoration, amende = 0, 0, 0
    if hors_delai:
        a_pct  = get_param(module, 'AMENDE_NON_DECLARATION', 15)
        amende = max(round(principal * a_pct / 100, 2), 500)
    if date_ech and date_decl > date_ech:
        penalite, majoration = calculer_penalites(principal, date_ech, date_decl, module)
    total = round(principal + penalite + majoration + amende, 2)
    if total < 200:
        total = 0; statut = 'sous_seuil'
    else:
        statut = 'emis'
    num = gen_num('DCL', 'declarations')
    conn = get_db()
    conn.execute('''INSERT INTO declarations
        (numero,module,reference_id,contribuable_id,commune_id,annee,trimestre,
         base_calcul,taux,montant_principal,penalite_retard,majoration,amende_non_declaration,montant_total,
         statut,date_declaration,date_echeance,agent_id,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (num, module, ref_id, contrib_id, 1, annee, int(f.get('trimestre', 0)),
         base, taux, principal, penalite, majoration, amende, total, statut,
         date_decl, date_ech, user['id'], f.get('notes', '')))
    conn.commit(); conn.close()
    flash(f'Déclaration {num} — Total: {total:.2f} DH ✅', 'success')
    routes_map = {
        'TNB':               'tnb.tnb_paiement',
        'DEBITS_BOISSONS':   'tdb.tdb_paiement',
        'STATIONNEMENT':     'sta.sta_paiement',
        'OCCUPATION_DOMAINE':'odp.odp_paiement',
        'FOURRIERE':         'fou.fou_paiement',
        'LOCATION_LOCAUX':   'loc.loc_paiement',
        'AFFERMAGE_SOUKS':   'sou.sou_paiement',
    }
    if module in routes_map:
        return redirect(url_for(routes_map[module], id=ref_id))
    return redirect(url_for('paiements'))


# ════════════════════════════════════════════════════════════
#  BULLETINS / PAIEMENTS
# ════════════════════════════════════════════════════════════
@app.route('/paiements')
@login_required
def paiements():
    user = get_current_user()
    conn = get_db()
    items = conn.execute('''SELECT b.*, d.module, d.annee, d.reference_id,
            c.nom, c.prenom, c.raison_sociale
        FROM bulletins b
        JOIN declarations d ON b.declaration_id=d.id
        LEFT JOIN contribuables c ON b.contribuable_id=c.id
        ORDER BY b.date_creation DESC''').fetchall()
    decls_sans_bulletin = conn.execute('''SELECT d.*, c.nom, c.prenom, c.raison_sociale FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.statut="emis" AND d.montant_total>0
        AND d.id NOT IN (SELECT declaration_id FROM bulletins WHERE statut IN ("en_attente","paye"))
        ORDER BY d.date_creation DESC''').fetchall()
    conn.close()
    return render_template('paiements/paiements.html', user=user, items=items,
                           decls=decls_sans_bulletin, today=date.today().isoformat())

@app.route('/bulletins/creer', methods=['POST'])
@login_required
def creer_bulletin():
    user = get_current_user()
    if not user['peut_creer_bulletin']:
        flash('Accès refusé', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    conn = get_db()
    decl = conn.execute('SELECT * FROM declarations WHERE id=?', (f['declaration_id'],)).fetchone()
    if decl:
        num = gen_num('BUL', 'bulletins', 'numero_bulletin', db_conn=conn)
        conn.execute('''INSERT INTO bulletins (numero_bulletin,declaration_id,contribuable_id,commune_id,montant,mode_paiement,date_paiement,agent_id,notes)
            VALUES (?,?,?,?,?,?,?,?,?)''',
            (num, decl['id'], decl['contribuable_id'], decl['commune_id'], decl['montant_total'],
             f.get('mode_paiement','especes'), f.get('date_paiement', date.today().isoformat()),
             user['id'], f.get('notes','')))
        conn.commit()
        flash(f'Bulletin {num} créé — En attente validation régisseur ✅', 'success')
    conn.close()
    return redirect(url_for('paiements'))

@app.route('/bulletins/<int:id>/valider', methods=['POST'])
@login_required
def valider_bulletin(id):
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé — Réservé au Régisseur', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    num_quittance = f.get('numero_quittance', '').strip()
    date_quittance = f.get('date_quittance', date.today().isoformat())
    if not num_quittance:
        flash('Le numéro de quittance est obligatoire', 'danger')
        return redirect(url_for('paiements'))
    conn = get_db()
    b = conn.execute('SELECT * FROM bulletins WHERE id=?', (id,)).fetchone()
    if b:
        # Valider CE bulletin
        conn.execute("""UPDATE bulletins 
            SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=?
            WHERE id=?""", (user['id'], num_quittance, date_quittance, id))
        conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                     (date_quittance, b['declaration_id']))
        # Valider aussi tous les autres bulletins avec le même numero_bulletin
        # (cas multi-trimestres TDB créés avec le même N° de BV)
        autres = conn.execute(
            "SELECT id, declaration_id FROM bulletins WHERE numero_bulletin=? AND id!=? AND statut='en_attente'",
            (b['numero_bulletin'], id)).fetchall()
        for ab in autres:
            conn.execute("UPDATE bulletins SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=? WHERE id=?",
                         (user['id'], num_quittance, date_quittance, ab['id']))
            conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                         (date_quittance, ab['declaration_id']))
        conn.commit()
        total_val = 1 + len(autres)
        # ── Si module FOURRIERE : passer le dossier en 'en_attente_sortie' ──
        decl_row = conn.execute('SELECT module, reference_id FROM declarations WHERE id=?',
                                (b['declaration_id'],)).fetchone()
        if decl_row and decl_row['module'] == 'FOURRIERE' and decl_row['reference_id']:
            conn.execute("UPDATE dossiers_fourriere SET statut='en_attente_sortie' WHERE id=?",
                         (decl_row['reference_id'],))
            conn.commit()
            flash(f'🚗 Dossier fourrière passé en "Attente sortie"', 'info')
        flash(f'✅ Paiement validé — Quittance N° {num_quittance} — {total_val} déclaration(s) soldée(s)', 'success')
    conn.close()
    return redirect(url_for('paiements'))


@app.route('/bulletins/<int:id>/rejeter', methods=['POST'])
@login_required
def rejeter_bulletin(id):
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    motif = f.get('motif_rejet', 'Non précisé').strip()
    conn = get_db()
    b = conn.execute('SELECT * FROM bulletins WHERE id=?', (id,)).fetchone()
    if b:
        conn.execute("UPDATE bulletins SET statut='rejete', motif_rejet=?, regisseur_id=? WHERE id=?",
                     (motif, user['id'], id))
        conn.execute("UPDATE declarations SET statut='emis' WHERE id=?", (b['declaration_id'],))
        conn.commit()
        flash(f'❌ Bulletin N° {b["numero_bulletin"]} rejeté : {motif}', 'danger')
    conn.close()
    return redirect(url_for('paiements'))

@app.route('/bulletins/valider-masse', methods=['POST'])
@login_required
def valider_bulletins_masse():
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé — Réservé au Régisseur', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    num_quittance = f.get('numero_quittance', '').strip()
    date_quittance = f.get('date_quittance', date.today().isoformat())
    bulletin_ids = f.getlist('bulletin_ids')
    if not num_quittance:
        flash('Le numéro de quittance est obligatoire', 'danger')
        return redirect(url_for('paiements'))
    if not bulletin_ids:
        flash('Aucun bulletin sélectionné', 'warn')
        return redirect(url_for('paiements'))
    conn = get_db()
    count = 0
    for bid in bulletin_ids:
        b = conn.execute("SELECT * FROM bulletins WHERE id=? AND statut='en_attente'", (bid,)).fetchone()
        if b:
            conn.execute("""UPDATE bulletins 
                SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=?
                WHERE id=?""", (user['id'], num_quittance, date_quittance, int(bid)))
            conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                         (date_quittance, b['declaration_id']))
            count += 1
    conn.commit()
    conn.close()
    flash(f'✅ {count} bulletin(s) validé(s) — Quittance N° {num_quittance}', 'success')
    return redirect(url_for('paiements'))


# ════════════════════════════════════════════════════════════
#  AVIS & RECOUVREMENT  (dashboard complet)
# ════════════════════════════════════════════════════════════

MODULE_LABELS = {
    'TNB': 'Taxe Terrains Non Bâtis',
    'DEBITS_BOISSONS': 'Débits de Boissons',
    'STATIONNEMENT': 'Stationnement TPV',
    'OCCUPATION_DOMAINE': 'Occupation Domaine Public',
    'FOURRIERE': 'Fourrière',
    'LOCATION_LOCAUX': 'Location Locaux Commerciaux',
    'AFFERMAGE_SOUKS': 'Affermage Souks',
}
MODULE_ICONS = {
    'TNB': '🏗️', 'DEBITS_BOISSONS': '🍺', 'STATIONNEMENT': '🚌',
    'OCCUPATION_DOMAINE': '🏕️', 'FOURRIERE': '🚗',
    'LOCATION_LOCAUX': '🏪', 'AFFERMAGE_SOUKS': '🛒',
}

def _get_impaye_par_module(conn, module=None):
    """Retourne tous les redevables avec arriérés, groupés par module."""
    q = '''
        SELECT d.id as decl_id, d.module, d.annee, d.trimestre,
               d.montant_total, d.montant_principal, d.penalite_retard,
               d.statut, d.date_echeance, d.numero as decl_numero,
               c.id as ctb_id, c.nom, c.prenom, c.raison_sociale,
               c.cin, c.ice, c.telephone, c.adresse, c.numero as ctb_numero,
               a.id as avis_id, a.numero_avis, a.date_emission as avis_date,
               a.statut as avis_statut, a.lot_id, a.lettre_id
        FROM declarations d
        JOIN contribuables c ON c.id = d.contribuable_id
        LEFT JOIN avis_non_paiement a ON a.declaration_id = d.id
        WHERE d.statut NOT IN ("paye","annule") AND d.montant_total > 0
    '''
    params = []
    if module:
        q += ' AND d.module = ?'
        params.append(module)
    q += ' ORDER BY d.module, c.nom, d.annee DESC'
    return conn.execute(q, params).fetchall()

def _get_stats_recouvrement(conn):
    """Stats de recouvrement par module et par année."""
    par_module = conn.execute('''
        SELECT module,
               COUNT(*) as nb_impayes,
               COUNT(DISTINCT contribuable_id) as nb_redevables,
               SUM(montant_total) as total_du,
               COUNT(CASE WHEN statut="emis" THEN 1 END) as nb_emis,
               SUM(CASE WHEN statut="paye" THEN montant_total ELSE 0 END) as total_recouvre
        FROM declarations
        GROUP BY module ORDER BY total_du DESC
    ''').fetchall()
    par_annee = conn.execute('''
        SELECT annee, module,
               COUNT(CASE WHEN statut NOT IN ("paye","annule") THEN 1 END) as nb_impayes,
               SUM(CASE WHEN statut NOT IN ("paye","annule") THEN montant_total ELSE 0 END) as total_du
        FROM declarations
        GROUP BY annee, module ORDER BY annee DESC, module
    ''').fetchall()
    return [dict(r) for r in par_module], [dict(r) for r in par_annee]

@app.route('/avis')
@login_required
def avis():
    user = get_current_user()
    conn = get_db()
    module_filtre = request.args.get('module', '')
    statut_filtre = request.args.get('statut', '')
    annee_filtre  = request.args.get('annee', '')

    # Impayés bruts
    rows = _get_impaye_par_module(conn, module_filtre or None)

    # Appliquer filtres statut/annee
    if statut_filtre:
        rows = [r for r in rows if (r['avis_statut'] or '') == statut_filtre]
    if annee_filtre:
        rows = [r for r in rows if str(r['annee']) == annee_filtre]

    # Grouper par module
    from collections import defaultdict
    modules_data = defaultdict(lambda: {'rows': [], 'total': 0, 'nb_ctb': set()})
    for r in rows:
        m = r['module']
        modules_data[m]['rows'].append(dict(r))
        modules_data[m]['total'] += float(r['montant_total'] or 0)
        modules_data[m]['nb_ctb'].add(r['ctb_id'])

    modules_list = []
    for mod, d in sorted(modules_data.items()):
        modules_list.append({
            'module': mod,
            'label': MODULE_LABELS.get(mod, mod),
            'icon': MODULE_ICONS.get(mod, '📋'),
            'rows': d['rows'],
            'total': round(d['total'], 2),
            'nb_redevables': len(d['nb_ctb']),
            'nb_avis_emis': sum(1 for r in d['rows'] if r.get('avis_id')),
        })

    # Stats globales
    stats_module, stats_annee = _get_stats_recouvrement(conn)
    total_global = sum(m['total'] for m in modules_list)

    # Historique lettres
    lettres = conn.execute('''
        SELECT l.*, u.nom as agent_nom
        FROM lettres_notification l
        LEFT JOIN utilisateurs u ON u.id = l.agent_id
        ORDER BY l.date_creation DESC LIMIT 50
    ''').fetchall()

    # Années disponibles pour filtre
    annees = [r['annee'] for r in conn.execute(
        'SELECT DISTINCT annee FROM declarations ORDER BY annee DESC').fetchall()]

    conn.close()
    return render_template('admin/avis.html',
        user=user, modules_list=modules_list,
        stats_module=stats_module, stats_annee=stats_annee,
        total_global=total_global, lettres=lettres,
        annees=annees, module_filtre=module_filtre,
        statut_filtre=statut_filtre, annee_filtre=annee_filtre,
        MODULE_LABELS=MODULE_LABELS, MODULE_ICONS=MODULE_ICONS,
        today=date.today().isoformat())



@app.route('/avis/generer', methods=['POST'])
@login_required
def generer_avis():
    user = get_current_user()
    conn = get_db()
    mode    = request.form.get('mode', 'lot')
    module  = request.form.get('module', '')
    lot_id  = f"LOT{datetime.now().strftime('%Y%m%d%H%M%S')}"

    q = '''SELECT * FROM declarations
           WHERE statut NOT IN ("paye","annule") AND montant_total > 0
           AND id NOT IN (SELECT declaration_id FROM avis_non_paiement WHERE statut="emis")'''
    params = []
    if module:
        q += ' AND module=?'
        params.append(module)

    decls = conn.execute(q, params).fetchall()
    count = 0
    for d in decls:
        num = gen_num('AVS', 'avis_non_paiement', 'numero_avis', db_conn=conn)
        conn.execute('''INSERT INTO avis_non_paiement
            (numero_avis,declaration_id,contribuable_id,commune_id,montant_du,date_emission,lot_id)
            VALUES (?,?,?,?,?,?,?)''',
            (num, d['id'], d['contribuable_id'], d['commune_id'],
             d['montant_total'], date.today().isoformat(), lot_id))
        count += 1

    conn.commit(); conn.close()
    flash(f'{count} avis générés ✅ (lot: {lot_id})', 'success')
    return redirect(url_for('avis'))


@app.route('/avis/generer-individuel', methods=['POST'])
@login_required
def generer_avis_individuel():
    decl_id = request.form.get('declaration_id')
    if not decl_id:
        flash('Déclaration manquante', 'danger')
        return redirect(url_for('avis'))
    conn = get_db()
    d = conn.execute('SELECT * FROM declarations WHERE id=?', (decl_id,)).fetchone()
    if d:
        num = gen_num('AVS', 'avis_non_paiement', 'numero_avis', db_conn=conn)
        lot_id = f"LOT{datetime.now().strftime('%Y%m%d%H%M%S')}"
        conn.execute('''INSERT INTO avis_non_paiement
            (numero_avis,declaration_id,contribuable_id,commune_id,montant_du,date_emission,lot_id)
            VALUES (?,?,?,?,?,?,?)''',
            (num, d['id'], d['contribuable_id'], d['commune_id'],
             d['montant_total'], date.today().isoformat(), lot_id))
        conn.commit()
        flash(f'Avis {num} créé ✅', 'success')
    conn.close()
    return redirect(url_for('avis'))


@app.route('/avis/lettre/generer', methods=['POST'])
@login_required
def generer_lettre():
    """Génère une lettre de notification en lot pour un module."""
    user = get_current_user()
    conn = get_db()
    module = request.form.get('module', '')
    type_lettre = request.form.get('type_lettre', 'relance')
    avis_ids_raw = request.form.get('avis_ids', '')
    avis_ids = [int(x) for x in avis_ids_raw.split(',') if x.strip().isdigit()]

    if not avis_ids:
        # Prendre tous les avis emis du module
        q = '''SELECT a.id, a.montant_du, a.contribuable_id, a.declaration_id
               FROM avis_non_paiement a
               JOIN declarations d ON d.id = a.declaration_id
               WHERE a.statut="emis"'''
        params = []
        if module:
            q += ' AND d.module=?'
            params.append(module)
        avis_rows = conn.execute(q, params).fetchall()
    else:
        avis_rows = conn.execute(
            f'SELECT id, montant_du, contribuable_id, declaration_id FROM avis_non_paiement WHERE id IN ({",".join("?" for _ in avis_ids)})',
            avis_ids
        ).fetchall()

    if not avis_rows:
        flash('Aucun avis à inclure dans la lettre.', 'warning')
        conn.close()
        return redirect(url_for('avis'))

    # Créer la lettre
    n = conn.execute('SELECT COUNT(*) as c FROM lettres_notification').fetchone()['c'] + 1
    num_lettre = f"LTR{datetime.now().year}{n:05d}"
    lot_id = f"LOT{datetime.now().strftime('%Y%m%d%H%M%S')}"
    total = sum(float(r['montant_du']) for r in avis_rows)

    conn.execute('''INSERT INTO lettres_notification
        (numero_lettre, lot_id, module, type_lettre, statut, date_generation,
         agent_id, nb_redevables, montant_total)
        VALUES (?,?,?,?,"brouillon",?,?,?,?)''',
        (num_lettre, lot_id, module, type_lettre, date.today().isoformat(),
         user['id'] if user else None, len(avis_rows), round(total, 2)))
    lettre_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

    for r in avis_rows:
        conn.execute('''INSERT INTO lettres_details
            (lettre_id, avis_id, declaration_id, contribuable_id, montant_du)
            VALUES (?,?,?,?,?)''',
            (lettre_id, r['id'], r['declaration_id'], r['contribuable_id'], r['montant_du']))
        conn.execute('UPDATE avis_non_paiement SET lettre_id=? WHERE id=?', (lettre_id, r['id']))

    conn.commit()
    flash(f'Lettre {num_lettre} créée en brouillon — {len(avis_rows)} redevable(s) ✅', 'success')
    conn.close()
    return redirect(url_for('lettre_imprimer', id=lettre_id))


@app.route('/avis/lettre/<int:id>/imprimer')
@login_required
def lettre_imprimer(id):
    """Page d'impression de la lettre (brouillon ou finale)."""
    user = get_current_user()
    conn = get_db()
    lettre = conn.execute('SELECT * FROM lettres_notification WHERE id=?', (id,)).fetchone()
    if not lettre:
        flash('Lettre introuvable', 'danger')
        conn.close()
        return redirect(url_for('avis'))

    details = conn.execute('''
        SELECT ld.*, c.nom, c.prenom, c.raison_sociale, c.cin, c.adresse, c.telephone,
               d.annee, d.module, d.numero as decl_numero, d.montant_total,
               d.montant_principal, d.penalite_retard,
               a.numero_avis, a.date_emission
        FROM lettres_details ld
        JOIN contribuables c ON c.id = ld.contribuable_id
        JOIN declarations d ON d.id = ld.declaration_id
        LEFT JOIN avis_non_paiement a ON a.id = ld.avis_id
        WHERE ld.lettre_id = ?
        ORDER BY c.nom, d.annee
    ''', (id,)).fetchall()

    commune = conn.execute('SELECT * FROM communes LIMIT 1').fetchone()
    conn.close()
    return render_template('admin/lettre_notification.html',
        user=user, lettre=dict(lettre),
        details=[dict(d) for d in details],
        commune=dict(commune) if commune else {},
        MODULE_LABELS=MODULE_LABELS,
        today=date.today().isoformat())


@app.route('/avis/lettre/<int:id>/approuver', methods=['POST'])
@login_required
def lettre_approuver(id):
    """Marque la lettre comme envoyée (approuvée)."""
    conn = get_db()
    conn.execute('''UPDATE lettres_notification
        SET statut="envoyee", date_envoi=? WHERE id=? AND statut="brouillon"''',
        (date.today().isoformat(), id))
    conn.commit()
    lettre = conn.execute('SELECT numero_lettre FROM lettres_notification WHERE id=?', (id,)).fetchone()
    conn.close()
    if lettre:
        flash(f'Lettre {lettre["numero_lettre"]} marquée comme envoyée ✅', 'success')
    return redirect(url_for('avis') + '#lettres')


@app.route('/avis/lettre/<int:id>/annuler', methods=['POST'])
@login_required
def lettre_annuler(id):
    conn = get_db()
    conn.execute('UPDATE lettres_notification SET statut="annulee" WHERE id=?', (id,))
    conn.execute('UPDATE avis_non_paiement SET lettre_id=NULL WHERE lettre_id=?', (id,))
    conn.commit(); conn.close()
    flash('Lettre annulée', 'warning')
    return redirect(url_for('avis') + '#lettres')


@app.route('/avis/export-excel')
@login_required
def avis_export_excel():
    """Export JSON des arriérés (utilisé par le JS pour générer Excel côté client)."""
    conn = get_db()
    module = request.args.get('module', '')
    rows = _get_impaye_par_module(conn, module or None)
    data = []
    for r in rows:
        data.append({
            'Module': MODULE_LABELS.get(r['module'], r['module']),
            'N° Déclaration': r['decl_numero'],
            'Redevable': f"{r['nom']} {r['prenom'] or r['raison_sociale'] or ''}".strip(),
            'CIN/ICE': r['cin'] or '',
            'Téléphone': r['telephone'] or '',
            'Adresse': r['adresse'] or '',
            'Année': r['annee'],
            'Montant Dû (DH)': round(float(r['montant_total'] or 0), 2),
            'N° Avis': r['numero_avis'] or '',
            'Date Avis': r['avis_date'] or '',
            'Statut Avis': r['avis_statut'] or 'Sans avis',
        })
    conn.close()
    return jsonify(data)



# ════════════════════════════════════════════════════════════
#  UTILISATEURS & COMMUNES
# ════════════════════════════════════════════════════════════
@app.route('/utilisateurs')
@login_required
def utilisateurs():
    user = get_current_user()
    conn = get_db()
    items   = conn.execute('''SELECT u.*, r.nom as role_nom FROM utilisateurs u
        JOIN roles r ON u.role_id=r.id WHERE u.actif=1''').fetchall()
    roles   = conn.execute('SELECT * FROM roles').fetchall()
    communes = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    # ── Données RBAC (onglet Droits par Module) ──────────────
    modules = conn.execute('SELECT * FROM app_modules WHERE actif=1 ORDER BY ordre').fetchall()
    perms_raw    = conn.execute('SELECT * FROM role_module_permissions').fetchall()
    perms = {}
    for p in perms_raw:
        perms.setdefault(p['role_id'], {})[p['module_code']] = dict(p)
    conn.close()
    return render_template('admin/utilisateurs.html',
        user=user, items=items, roles=roles, communes=communes,
        modules=modules, perms=perms)

@app.route('/utilisateurs/ajouter', methods=['POST'])
@login_required
def ajouter_utilisateur():
    f = request.form
    pwd = hashlib.sha256(f['password'].encode()).hexdigest()
    conn = get_db()
    existing = conn.execute('SELECT id FROM utilisateurs WHERE email=?', (f['email'],)).fetchone()
    if existing:
        flash('Cet email est déjà utilisé ❌', 'danger')
    else:
        conn.execute('INSERT INTO utilisateurs (nom,prenom,email,mot_de_passe,role_id,commune_id) VALUES (?,?,?,?,?,?)',
            (f['nom'], f['prenom'], f['email'], pwd, f['role_id'], f.get('commune_id') or 1))
        conn.commit()
        flash('Utilisateur ajouté ✅', 'success')
    conn.close()
    return redirect(url_for('utilisateurs'))

@app.route('/utilisateurs/<int:id>/modifier', methods=['POST'])
@login_required
def modifier_utilisateur(id):
    f = request.form
    conn = get_db()
    if f.get('password'):
        pwd = hashlib.sha256(f['password'].encode()).hexdigest()
        conn.execute('UPDATE utilisateurs SET nom=?,prenom=?,email=?,mot_de_passe=?,role_id=?,commune_id=? WHERE id=?',
            (f['nom'], f['prenom'], f['email'], pwd, f['role_id'], f.get('commune_id') or 1, id))
    else:
        conn.execute('UPDATE utilisateurs SET nom=?,prenom=?,email=?,role_id=?,commune_id=? WHERE id=?',
            (f['nom'], f['prenom'], f['email'], f['role_id'], f.get('commune_id') or 1, id))
    conn.commit(); conn.close()
    flash('Utilisateur modifié ✅', 'success')
    return redirect(url_for('utilisateurs'))

@app.route('/utilisateurs/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_utilisateur(id):
    conn = get_db()
    # Empêcher suppression de son propre compte
    if id == session.get('user_id'):
        flash('Impossible de supprimer votre propre compte ❌', 'danger')
        conn.close()
        return redirect(url_for('utilisateurs'))
    conn.execute('UPDATE utilisateurs SET actif=0 WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Utilisateur désactivé ✅', 'success')
    return redirect(url_for('utilisateurs'))

# ── Rôles ────────────────────────────────────────────────────
@app.route('/roles/ajouter', methods=['POST'])
@login_required
def ajouter_role():
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès refusé ❌', 'danger')
        return redirect(url_for('utilisateurs'))
    f = request.form
    conn = get_db()
    conn.execute('''INSERT OR IGNORE INTO roles
        (nom,peut_voir,peut_ajouter,peut_modifier,peut_supprimer,peut_creer_bulletin,peut_valider_paiement,peut_config)
        VALUES (?,?,?,?,?,?,?,?)''',
        (f['nom'],
         1 if f.get('peut_voir') else 0,
         1 if f.get('peut_ajouter') else 0,
         1 if f.get('peut_modifier') else 0,
         1 if f.get('peut_supprimer') else 0,
         1 if f.get('peut_creer_bulletin') else 0,
         1 if f.get('peut_valider_paiement') else 0,
         1 if f.get('peut_config') else 0))
    conn.commit(); conn.close()
    flash('Rôle créé ✅', 'success')
    return redirect(url_for('utilisateurs'))

@app.route('/roles/<int:id>/modifier', methods=['POST'])
@login_required
def modifier_role(id):
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès refusé ❌', 'danger')
        return redirect(url_for('utilisateurs'))
    f = request.form
    conn = get_db()
    conn.execute('''UPDATE roles SET nom=?,peut_voir=?,peut_ajouter=?,peut_modifier=?,
        peut_supprimer=?,peut_creer_bulletin=?,peut_valider_paiement=?,peut_config=? WHERE id=?''',
        (f['nom'],
         1 if f.get('peut_voir') else 0,
         1 if f.get('peut_ajouter') else 0,
         1 if f.get('peut_modifier') else 0,
         1 if f.get('peut_supprimer') else 0,
         1 if f.get('peut_creer_bulletin') else 0,
         1 if f.get('peut_valider_paiement') else 0,
         1 if f.get('peut_config') else 0,
         id))
    conn.commit(); conn.close()
    flash('Rôle modifié ✅', 'success')
    return redirect(url_for('utilisateurs'))

@app.route('/roles/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_role(id):
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès refusé ❌', 'danger')
        return redirect(url_for('utilisateurs'))
    conn = get_db()
    # Vérifier qu'aucun utilisateur n'utilise ce rôle
    nb = conn.execute('SELECT COUNT(*) FROM utilisateurs WHERE role_id=? AND actif=1', (id,)).fetchone()[0]
    if nb > 0:
        flash(f'Impossible : {nb} utilisateur(s) utilisent ce rôle ❌', 'danger')
    else:
        conn.execute('DELETE FROM roles WHERE id=?', (id,))
        conn.commit()
        flash('Rôle supprimé ✅', 'success')
    conn.close()
    return redirect(url_for('utilisateurs'))


# ════════════════════════════════════════════════════════════
#  GESTION DROITS PAR MODULE (RBAC)
# ════════════════════════════════════════════════════════════

@app.route('/roles/permissions')
@login_required
def roles_permissions():
    """Redirige vers la page unifiée Utilisateurs & Rôles, onglet Droits."""
    return redirect(url_for('utilisateurs') + '#tab-droits')


@app.route('/roles/<int:role_id>/permissions/sauvegarder', methods=['POST'])
@login_required
def sauvegarder_permissions_role(role_id):
    """Sauvegarde les permissions par module pour un rôle."""
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès refusé ❌', 'danger')
        return redirect(url_for('roles_permissions'))
    conn = get_db()
    modules = conn.execute('SELECT code FROM app_modules WHERE actif=1').fetchall()
    for m in modules:
        code = m['code']
        voir      = 1 if request.form.get(f'{code}_voir')      else 0
        ajouter   = 1 if request.form.get(f'{code}_ajouter')   else 0
        modifier  = 1 if request.form.get(f'{code}_modifier')  else 0
        supprimer = 1 if request.form.get(f'{code}_supprimer') else 0
        conn.execute('''INSERT INTO role_module_permissions
            (role_id, module_code, peut_voir, peut_ajouter, peut_modifier, peut_supprimer)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(role_id, module_code) DO UPDATE SET
                peut_voir=excluded.peut_voir,
                peut_ajouter=excluded.peut_ajouter,
                peut_modifier=excluded.peut_modifier,
                peut_supprimer=excluded.peut_supprimer''',
            (role_id, code, voir, ajouter, modifier, supprimer))
    conn.commit()
    conn.close()
    flash('Permissions mises à jour ✅', 'success')
    return redirect(url_for('utilisateurs') + '#tab-droits')


@app.route('/api/roles/<int:role_id>/permissions')
@login_required
def api_role_permissions(role_id):
    """Retourne les permissions d'un rôle en JSON (pour AJAX)."""
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM role_module_permissions WHERE role_id=?', (role_id,)
    ).fetchall()
    conn.close()
    return jsonify({r['module_code']: dict(r) for r in rows})


@app.route('/utilisateurs/<int:user_id>/permissions')
@login_required
def user_permissions(user_id):
    """Affiche les permissions effectives d'un utilisateur."""
    current = get_current_user()
    if not current['peut_config']:
        flash('Accès refusé ❌', 'danger')
        return redirect(url_for('utilisateurs'))
    conn = get_db()
    target_user = conn.execute(
        '''SELECT u.*, r.nom as role_nom, r.peut_voir, r.peut_ajouter,
           r.peut_modifier, r.peut_supprimer, r.peut_config,
           r.peut_creer_bulletin, r.peut_valider_paiement
           FROM utilisateurs u JOIN roles r ON u.role_id=r.id WHERE u.id=?''',
        (user_id,)).fetchone()
    if not target_user:
        flash('Utilisateur introuvable', 'danger')
        conn.close()
        return redirect(url_for('utilisateurs'))
    modules = conn.execute('SELECT * FROM app_modules WHERE actif=1 ORDER BY ordre').fetchall()
    perms_raw = conn.execute(
        'SELECT * FROM role_module_permissions WHERE role_id=?',
        (target_user['role_id'],)).fetchall()
    perms = {p['module_code']: dict(p) for p in perms_raw}
    conn.close()
    return render_template('admin/user_permissions.html',
        user=current, target_user=target_user,
        modules=modules, perms=perms)


@app.route('/communes')
@login_required
def communes():
    """Redirige vers l'onglet Commune de Paramètres Système."""
    return redirect(url_for('parametres_systeme'))


@app.route('/communes/modifier', methods=['POST'])
@login_required
def modifier_commune():
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès refusé', 'danger')
        return redirect(url_for('index'))

    f = request.form
    file = request.files.get('logo')
    logo_path = None
    if file and file.filename != '':
        target_dir = os.path.join(app.static_folder, 'img')
        os.makedirs(target_dir, exist_ok=True)
        try:
            from PIL import Image
            img = Image.open(file.stream)
            img.save(os.path.join(target_dir, 'logo.png'), format='PNG')
        except Exception:
            file.seek(0)
            file.save(os.path.join(target_dir, 'logo.png'))
        logo_path = 'img/logo.png'

    conn = get_db()
    c = conn.execute('SELECT id, logo FROM communes LIMIT 1').fetchone()
    logo_path = logo_path or (c['logo'] if c else None)

    if c:
        conn.execute('''UPDATE communes SET 
            nom=?, nom_ar=?, president_fr=?, president_ar=?,
            region=?, region_ar=?, province=?, province_ar=?, logo=?
            WHERE id=?''',
            (f['nom'], f.get('nom_ar',''), f.get('president_fr',''), f.get('president_ar',''),
             f.get('region',''), f.get('region_ar',''), f.get('province',''), f.get('province_ar',''), 
             logo_path, c['id']))
    else:
        conn.execute('''INSERT INTO communes 
            (nom, nom_ar, president_fr, president_ar, region, region_ar, province, province_ar, logo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (f['nom'], f.get('nom_ar',''), f.get('president_fr',''), f.get('president_ar',''),
             f.get('region',''), f.get('region_ar',''), f.get('province',''), f.get('province_ar',''), 
             logo_path))
    conn.commit()
    conn.close()

    # ── Synchronisation silencieuse config.json ──────────────
    # Le launcher lit config.json → on y recopie les infos DB automatiquement
    try:
        cfg = _load_sys_config()
        cfg['commune'] = {
            'nom':      f.get('nom', ''),
            'nom_ar':   f.get('nom_ar', ''),
            'region':   f.get('region', ''),
            'region_ar':f.get('region_ar', ''),
            'province': f.get('province', ''),
            'province_ar':f.get('province_ar', ''),
            'logo':     logo_path,
        }
        _save_sys_config(cfg)
    except Exception as e:
        logger.warning(f"Sync config.json: {e}")

    flash('Informations de la commune mises à jour ✅', 'success')
    return redirect(url_for('parametres_systeme'))


# ════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ════════════════════════════════════════════════════════════
@app.route('/export/<module>/excel')
@login_required
def export_excel(module):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    conn = get_db()
    data = conn.execute('''SELECT d.numero, c.nom||" "||COALESCE(c.prenom,"") as contribuable, c.cin,
        d.annee, d.base_calcul, d.taux, d.montant_principal, d.penalite_retard, d.majoration,
        d.amende_non_declaration, d.montant_total, d.statut, d.date_declaration
        FROM declarations d JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.module=? ORDER BY d.date_creation DESC''', (module,)).fetchall()
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = module[:31]
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1e3a5f')
    hdrs = ['N° Décl.','Contribuable','CIN','Année','Base','Taux%','Principal',
            'Pénalité','Majoration','Amende','TOTAL','Statut','Date']
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hf; cell.fill = hfill
        ws.column_dimensions[cell.column_letter].width = 15
    for r, row in enumerate(data, 2):
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{module}_{datetime.now():%Y%m%d}.xlsx')


# ════════════════════════════════════════════════════════════
#  API JSON
# ════════════════════════════════════════════════════════════
@app.route('/api/contribuable/<int:id>')
@login_required
def api_contribuable(id):
    conn = get_db()
    c = conn.execute('SELECT * FROM contribuables WHERE id=?', (id,)).fetchone()
    conn.close()
    return jsonify(dict(c) if c else {})

@app.route('/api/calcul')
@login_required
def api_calcul():
    montant    = float(request.args.get('montant', 0))
    ech        = request.args.get('echeance', '')
    pay        = request.args.get('paiement', date.today().isoformat())
    module     = request.args.get('module', 'GLOBAL')
    hors_delai = request.args.get('hors_delai', '0') == '1'
    p, m = calculer_penalites(montant, ech, pay, module)
    amende = 0
    if hors_delai:
        a_pct  = get_param(module, 'AMENDE_NON_DECLARATION', 15)
        amende = max(round(montant * a_pct / 100, 2), 500)
    return jsonify({'penalite': p, 'majoration': m, 'amende': amende,
                    'total': round(montant + p + m + amende, 2)})

@app.route('/api/tarifs/<module>')
@login_required
def api_tarifs(module):
    conn = get_db()
    today = date.today().isoformat()
    tarifs = conn.execute('''SELECT t.* FROM tarifs t
        JOIN rubriques r ON t.rubrique_id=r.id
        WHERE r.module=? AND t.actif=1
          AND t.date_debut <= ?
          AND (t.date_fin IS NULL OR t.date_fin >= ?)
        ORDER BY t.libelle''', (module, today, today)).fetchall()
    conn.close()
    return jsonify([dict(t) for t in tarifs])

@app.route('/api/stats')
@login_required
def api_stats():
    conn = get_db()
    modules = ['TNB','DEBITS_BOISSONS','TRANSPORT_VOYAGEURS','STATIONNEMENT',
               'OCCUPATION_DOMAINE','FOURRIERE','LOCATION_LOCAUX','AFFERMAGE_SOUKS']
    result = {}
    for m in modules:
        r = conn.execute("SELECT COUNT(*) as c, COALESCE(SUM(montant_total),0) as t FROM declarations WHERE module=?", (m,)).fetchone()
        result[m] = {'count': r['c'], 'total': round(r['t'], 2)}
    conn.close()
    return jsonify(result)




# ════════════════════════════════════════════════════════════
#  CONFIGURATION SYSTÈME — config.json helpers
# ════════════════════════════════════════════════════════════
CONFIG_FILE   = 'config.json'
BACKUP_LOG    = 'backup_log.json'
VERSION_FILE  = 'version.txt'
GITHUB_USER   = 'Yomix90'
GITHUB_REPO   = 'JIBAYAT'
GITHUB_BRANCH = 'main'

ALL_MODULES = {
    'TNB':                 '🏗️  Taxe Terrains Non Bâtis',
    'DEBITS_BOISSONS':     '🍺  Débits de Boissons',
    'TRANSPORT_VOYAGEURS': '🚌  Transport Public Voyageurs',
    'STATIONNEMENT':       '🅿️  Stationnement TPV',
    'OCCUPATION_DOMAINE':  '🏪  Occupation Domaine Public',
    'FOURRIERE':           '🔒  Droits de Fourrière',
    'LOCATION_LOCAUX':     '🏢  Location Locaux Commerciaux',
    'AFFERMAGE_SOUKS':     '🛒  Affermage Souks Communaux',
}

def _load_sys_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def _save_sys_config(data: dict) -> None:
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _read_version() -> str:
    try:
        with open(VERSION_FILE, 'r') as f:
            return f.read().strip()
    except Exception:
        return '1.0.0'

def _load_backup_log() -> list:
    if os.path.exists(BACKUP_LOG):
        try:
            with open(BACKUP_LOG, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def _append_backup_log(entry: dict) -> None:
    logs = _load_backup_log()
    logs.insert(0, entry)
    with open(BACKUP_LOG, 'w', encoding='utf-8') as f:
        json.dump(logs[:50], f, ensure_ascii=False, indent=2)


# ════════════════════════════════════════════════════════════
#  ROUTE : PREMIÈRE INSTALLATION  /setup
# ════════════════════════════════════════════════════════════
@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """Page de première installation — accessible sans login si pas de config."""
    # Si config déjà présente et utilisateur connecté → paramètres
    if os.path.exists(CONFIG_FILE) and 'user_id' in session:
        return redirect(url_for('parametres_systeme'))

    error = None
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip()
        if not nom:
            error = "Le nom de la commune est obligatoire."
        else:
            logo_path = None
            file = request.files.get('logo')
            if file and file.filename != '':
                target_dir = os.path.join(app.static_folder, 'img')
                os.makedirs(target_dir, exist_ok=True)
                try:
                    from PIL import Image
                    img = Image.open(file.stream)
                    img.save(os.path.join(target_dir, 'logo.png'), format='PNG')
                except Exception:
                    file.seek(0)
                    file.save(os.path.join(target_dir, 'logo.png'))
                logo_path = 'img/logo.png'

            cfg = {
                'commune': {
                    'nom':      nom,
                    'nom_ar':   request.form.get('nom_ar', '').strip(),
                    'region':   request.form.get('region', '').strip(),
                    'region_ar':request.form.get('region_ar', '').strip(),
                    'province': request.form.get('province', '').strip(),
                    'province_ar':request.form.get('province_ar', '').strip(),
                },
                'modules': request.form.getlist('modules') or list(ALL_MODULES.keys()),
                'auto_backup': True,
            }
            if logo_path:
                cfg['commune']['logo'] = logo_path
            _save_sys_config(cfg)
            try:
                init_db()
            except Exception as e:
                error = f"Erreur initialisation BD : {e}"
                if os.path.exists(CONFIG_FILE):
                    os.remove(CONFIG_FILE)
            if not error:
                flash('✅ Installation réussie ! Bienvenue dans JIBAYAT.', 'success')
                return redirect(url_for('login'))

    return render_template('setup.html',
                           error=error,
                           all_modules=ALL_MODULES,
                           version=_read_version())


# ════════════════════════════════════════════════════════════
#  ROUTE : PARAMÈTRES SYSTÈME  /parametres-systeme
# ════════════════════════════════════════════════════════════
@app.route('/parametres-systeme')
@login_required
def parametres_systeme():
    user = get_current_user()
    if not user['peut_config']:
        flash('Accès réservé aux administrateurs.', 'danger')
        return redirect(url_for('index'))

    cfg      = _load_sys_config()
    logs     = _load_backup_log()
    version  = _read_version()
    db_size  = round(os.path.getsize('fiscalite.db') / 1024 / 1024, 2) if os.path.exists('fiscalite.db') else 0
    db_exists = os.path.exists('fiscalite.db')

    # Données commune depuis la base de données
    conn = get_db()
    commune_db = conn.execute('SELECT * FROM communes LIMIT 1').fetchone()
    conn.close()

    return render_template('parametres_systeme.html',
                           user=user,
                           cfg=cfg,
                           commune_db=commune_db,
                           all_modules=ALL_MODULES,
                           logs=logs,
                           version=version,
                           db_size=db_size,
                           db_exists=db_exists)


# ── API Paramètres Système ──────────────────────────────────

@app.route('/api/systeme/commune', methods=['POST'])
@login_required
def api_systeme_commune():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    nom = request.form.get('nom', '').strip()
    if not nom:
        return jsonify({'ok': False, 'error': 'Nom obligatoire'})

    cfg = _load_sys_config()
    cfg['commune'] = {
        'nom':      nom,
        'nom_ar':   request.form.get('nom_ar', '').strip(),
        'region':   request.form.get('region', '').strip(),
        'region_ar':request.form.get('region_ar', '').strip(),
        'province': request.form.get('province', '').strip(),
        'province_ar':request.form.get('province_ar', '').strip(),
    }
    _save_sys_config(cfg)
    return jsonify({'ok': True, 'msg': 'Informations commune sauvegardées.'})


@app.route('/api/systeme/modules', methods=['POST'])
@login_required
def api_systeme_modules():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    modules = request.form.getlist('modules')
    cfg = _load_sys_config()
    cfg['modules'] = modules
    _save_sys_config(cfg)
    return jsonify({'ok': True, 'msg': f'{len(modules)} module(s) activé(s).'})


@app.route('/api/systeme/backup-config', methods=['POST'])
@login_required
def api_systeme_backup_config():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    cfg = _load_sys_config()
    cfg['auto_backup']     = request.form.get('auto_backup') == '1'
    cfg['gdrive_backup']   = request.form.get('gdrive_backup', '').strip()
    cfg['gdrive_webhook']  = request.form.get('gdrive_webhook', '').strip()
    cfg['gdrive_folder_id'] = request.form.get('gdrive_folder_id', '').strip()
    _save_sys_config(cfg)
    return jsonify({'ok': True, 'msg': 'Configuration sauvegarde sauvegardée.'})


@app.route('/api/systeme/github-config', methods=['POST'])
@login_required
def api_systeme_github_config():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    cfg = _load_sys_config()
    cfg['github_token'] = request.form.get('github_token', '').strip()
    _save_sys_config(cfg)
    return jsonify({'ok': True, 'msg': 'Configuration GitHub sauvegardée.'})


@app.route('/api/systeme/feedback-config', methods=['POST'])
@login_required
def api_systeme_feedback_config():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    cfg = _load_sys_config()
    cfg['feedback_webhook'] = request.form.get('feedback_webhook', '').strip()
    _save_sys_config(cfg)
    return jsonify({'ok': True, 'msg': 'Configuration Feedback sauvegardée.'})


@app.route('/api/systeme/feedback', methods=['POST'])
@login_required
def api_systeme_feedback():
    user = get_current_user()
    cfg = _load_sys_config()
    webhook = cfg.get('feedback_webhook', '').strip()

    if not webhook:
        return jsonify({'ok': False, 'error': "Le webhook d'avis n'est pas configuré par l'administrateur."})

    type_dmd = request.form.get('type', 'Suggestion')
    message  = request.form.get('message', '').strip()
    page     = request.form.get('page', 'Général')
    
    if not message:
        return jsonify({'ok': False, 'error': "Le message ne peut pas être vide."})

    conn = get_db()
    commune_db = conn.execute('SELECT nom FROM communes LIMIT 1').fetchone()
    conn.close()
    commune_nom = commune_db['nom'] if commune_db and commune_db['nom'] else 'Inconnue'

    import requests as _req
    payload = {
        'type': type_dmd,
        'utilisateur': f"{user['nom'] or ''} {user['prenom'] or ''}".strip(),
        'commune': commune_nom,
        'page': page,
        'message': message
    }

    try:
        r = _req.post(webhook, json=payload, timeout=10, allow_redirects=True)
        if r.status_code == 200:
            return jsonify({'ok': True, 'msg': 'Votre message a été envoyé avec succès !'})
        else:
            return jsonify({'ok': False, 'error': f"Erreur serveur Google ({r.status_code})"})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/systeme/backup-now', methods=['POST'])
@login_required
def api_systeme_backup_now():
    """Déclenche une sauvegarde manuelle immédiate."""
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403

    if not os.path.exists('fiscalite.db'):
        return jsonify({'ok': False, 'error': 'Base de données introuvable.'})

    cfg      = _load_sys_config()
    dt       = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'fiscalite_Manuel_{dt}.db'
    results  = []

    # ── Option A : Dossier local ─────────────────────────────
    dest_dir = cfg.get('gdrive_backup', '')
    if dest_dir and os.path.exists(dest_dir):
        try:
            shutil.copy('fiscalite.db', os.path.join(dest_dir, filename))
            results.append(f'✅ Copie locale : {os.path.join(dest_dir, filename)}')
            _append_backup_log({'date': dt, 'type': 'Manuel Local',
                                'dest': dest_dir, 'status': '✅ Succès'})
        except Exception as ex:
            results.append(f'❌ Erreur locale : {ex}')
            _append_backup_log({'date': dt, 'type': 'Manuel Local',
                                'dest': dest_dir, 'status': f'❌ {ex}'})

    # ── Option B : Google Drive via Apps Script ───────────────
    webhook   = cfg.get('gdrive_webhook', '').strip()
    folder_id = cfg.get('gdrive_folder_id', '').strip()
    if webhook and folder_id:
        try:
            import requests as _req

            with open('fiscalite.db', 'rb') as f_db:
                file_b64 = base64.b64encode(f_db.read()).decode('utf-8')

            payload = {
                'filename':  filename,
                'folder_id': folder_id,
                'mimeType':  'application/x-sqlite3',
                'file':      file_b64,
            }

            # requests suit automatiquement les redirections 302 de Google
            r = _req.post(webhook, json=payload, timeout=60,
                          allow_redirects=True)

            if r.status_code == 200:
                try:
                    resp_json = r.json()
                    if resp_json.get('ok'):
                        results.append('✅ Sauvegarde Cloud Google Drive réussie.')
                        _append_backup_log({'date': dt, 'type': 'Manuel Cloud',
                                            'dest': 'Google Drive', 'status': '✅ Succès'})
                    else:
                        err = resp_json.get('error', r.text[:200])
                        results.append(f'❌ Google Drive : {err}')
                        _append_backup_log({'date': dt, 'type': 'Manuel Cloud',
                                            'dest': 'Google Drive', 'status': f'❌ {err}'})
                except Exception:
                    # Réponse non-JSON mais HTTP 200 → probablement ok
                    results.append('✅ Sauvegarde Cloud envoyée (réponse non-JSON).')
                    _append_backup_log({'date': dt, 'type': 'Manuel Cloud',
                                        'dest': 'Google Drive', 'status': '✅ Succès'})
            else:
                err = f'HTTP {r.status_code} — {r.text[:200]}'
                results.append(f'❌ Erreur Cloud : {err}')
                _append_backup_log({'date': dt, 'type': 'Manuel Cloud',
                                    'dest': 'Google Drive', 'status': f'❌ {err}'})

        except Exception as ex:
            results.append(f'❌ Erreur Cloud : {ex}')
            _append_backup_log({'date': dt, 'type': 'Manuel Cloud',
                                'dest': 'Google Drive', 'status': f'❌ {ex}'})

    # ── Aucune option configurée → copie de précaution locale ─
    if not results:
        try:
            shutil.copy('fiscalite.db', filename)
            results.append(f'✅ Copie de précaution créée : {filename}')
            _append_backup_log({'date': dt, 'type': 'Manuel (précaution)',
                                'dest': filename, 'status': '✅ Succès'})
        except Exception as ex:
            return jsonify({'ok': False, 'error': str(ex)})

    return jsonify({'ok': True, 'msg': '\n'.join(results), 'logs': _load_backup_log()[:10]})


@app.route('/api/systeme/init-db', methods=['POST'])
@login_required
def api_systeme_init_db():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403
    try:
        init_db()
        return jsonify({'ok': True, 'msg': 'Base de données initialisée/mise à jour avec succès.'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/systeme/test-gdrive', methods=['POST'])
@login_required
def api_systeme_test_gdrive():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403
    webhook   = request.form.get('webhook', '').strip()
    folder_id = request.form.get('folder_id', '').strip()
    if not webhook or not folder_id:
        return jsonify({'ok': False, 'error': 'Webhook et dossier obligatoires.'})
    try:
        import requests as _req
        payload = {'test': True, 'filename': 'jibayat_test.txt',
                   'folder_id': folder_id, 'mimeType': 'text/plain', 'file': 'dGVzdA=='}
        r = _req.post(webhook, json=payload, timeout=15, allow_redirects=True)
        if r.status_code == 200:
            return jsonify({'ok': True, 'msg': '✅ Connexion Google Drive réussie !'})
        return jsonify({'ok': False, 'error': f'HTTP {r.status_code}'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/systeme/check-update')
@login_required
def api_systeme_check_update():
    try:
        import requests as _req
        cfg = _load_sys_config()
        token = cfg.get('github_token', '').strip()
        
        # On utilise l'API GitHub plutôt que raw.githubusercontent pour supporter les tokens privés
        url = f'https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/version.txt'
        headers = {'Accept': 'application/vnd.github.v3.raw'}
        if token:
            headers['Authorization'] = f'token {token}'
            
        r = _req.get(url, headers=headers, timeout=10)
        
        if r.status_code == 200:
            remote = r.text.strip()
            local  = _read_version()
            def vt(v):
                try: return tuple(int(x) for x in v.split('.'))
                except: return (0,)
            has_update = vt(remote) > vt(local)
            return jsonify({'ok': True, 'local': local, 'remote': remote, 'has_update': has_update})
        elif r.status_code == 404:
            return jsonify({'ok': False, 'error': "Dépôt privé ou introuvable (404). Configurez un Token GitHub."})
        else:
            return jsonify({'ok': False, 'error': f'Erreur GitHub (HTTP {r.status_code})'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/systeme/do-update', methods=['POST'])
@login_required
def api_systeme_do_update():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403
    
    try:
        import subprocess, sys, os
        import urllib.request
        
        # Détecter si on est dans un .exe compilé par PyInstaller
        is_exe = getattr(sys, 'frozen', False)
        executable = sys.executable if is_exe else 'LANCER.bat'
        
        if is_exe:
            # Mode EXE : Télécharger l'archive ZIP depuis la dernière Release GitHub
            zip_url = f"https://github.com/{GITHUB_USER}/{GITHUB_REPO}/releases/latest/download/JIBAYAT-update.zip"
            zip_path = "JIBAYAT-update.zip"
            
            try:
                # Téléchargement bloquant (l'UI montre le loading)
                urllib.request.urlretrieve(zip_url, zip_path)
            except Exception as e:
                return jsonify({'ok': False, 'error': f"Impossible de télécharger la mise à jour (Release introuvable sur GitHub). Assurez-vous d'avoir publié une Release avec le fichier JIBAYAT-update.zip. Détails: {str(e)}"})
            
            # Script BATCH pour le mode EXE
            bat_content = f"""@echo off
:: Attendre 3 secondes pour que l'application actuelle se ferme complètement
ping 127.0.0.1 -n 4 > nul

:: Extraire le ZIP téléchargé en écrasant les fichiers existants
powershell -command "Expand-Archive -Path '{zip_path}' -DestinationPath '.' -Force"

:: Supprimer l'archive
del "{zip_path}"

:: Relancer l'application
start "" "{executable}"

:: Supprimer ce script temporaire
del "%~f0"
"""
        else:
            # Mode Source : Utiliser git pull
            bat_content = f"""@echo off
:: Attendre 3 secondes pour que l'application actuelle se ferme
ping 127.0.0.1 -n 4 > nul

:: Faire le git pull
git pull origin main

:: Relancer l'application
start "" "{executable}"

:: Supprimer ce script temporaire
del "%~f0"
"""

        with open('update_temp.bat', 'w', encoding='utf-8') as f:
            f.write(bat_content)

        # Lancer le script de mise à jour de façon cachée
        startupinfo = None
        if os.name == 'nt':
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
            
        subprocess.Popen(['update_temp.bat'], startupinfo=startupinfo)
        
        # Quitter immédiatement l'application pour libérer les fichiers
        import threading
        threading.Thread(target=lambda: os._exit(0)).start()
        
        return jsonify({'ok': True, 'msg': 'Mise à jour téléchargée avec succès. L\'application va redémarrer...'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/systeme/export-db')
@login_required
def api_systeme_export_db():
    if not os.path.exists('fiscalite.db'):
        return jsonify({'ok': False, 'error': 'Base introuvable'}), 404
    dt = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    return send_file('fiscalite.db',
                     as_attachment=True,
                     download_name=f'fiscalite_Export_{dt}.db',
                     mimetype='application/octet-stream')


@app.route('/api/systeme/import-db', methods=['POST'])
@login_required
def api_systeme_import_db():
    user = get_current_user()
    if not user['peut_config']:
        return jsonify({'ok': False, 'error': 'Accès refusé'}), 403
    if 'db_file' not in request.files:
        return jsonify({'ok': False, 'error': 'Aucun fichier reçu.'})
    f = request.files['db_file']
    if not f.filename.endswith('.db'):
        return jsonify({'ok': False, 'error': 'Seuls les fichiers .db sont acceptés.'})
    dt = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    try:
        if os.path.exists('fiscalite.db'):
            shutil.copy('fiscalite.db', f'fiscalite_AvantImport_{dt}.db')
        f.save('fiscalite.db')
        _append_backup_log({'date': dt, 'type': 'Restauration', 'dest': 'fiscalite.db', 'status': '✅ Succès'})
        return jsonify({'ok': True, 'msg': f'Base restaurée. Ancienne sauvegarde : fiscalite_AvantImport_{dt}.db'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/systeme/backup-logs')
@login_required
def api_systeme_backup_logs():
    return jsonify(_load_backup_log())


# ════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE
# ════════════════════════════════════════════════════════════
if __name__ == '__main__':
    init_db()
    import socket as _sock
    try:
        s = _sock.socket(_sock.AF_INET, _sock.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]; s.close()
    except:
        ip = 'localhost'
    print(f"\n{'='*55}\n  JIBAYAT — Gestion Fiscale Communale\n  Local : http://localhost:5050\n  Réseau: http://{ip}:5050\n  Login : admin@commune.ma / admin123\n{'='*55}\n")
    app.run(host='0.0.0.0', port=5050, debug=False)

